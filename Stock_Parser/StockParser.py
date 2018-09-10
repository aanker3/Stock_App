#!/usr/bin/env python3

# Name            :  StockParser
# Script Name     :  StockParser.py

import datetime
import re
from lxml import html
import requests
import time
import Tkinter
from PIL import ImageTk, Image
import os.path
import subprocess
import xlsxwriter #pip install xlsxwriter
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles.colors import RED
from openpyxl.styles.colors import GREEN
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter, column_index_from_string, coordinate_from_string
import copy
from openpyxl.worksheet import *

#google and morning star are not working for some reason.  List out of range.  has to do with the [0].text.  Not sure

#def GetGoogle_MainPage(ticker):
#https://www.google.com/search?tbm=fin&ei=2ntPW7LGGM-p_QbjpqyYAQ&q=goog&oq=goog&gs_l=finance-immersive.3..81l3.9658.10169.0.10352.4.4.0.0.0.0.225.225.2-1.1.0..1..0...1.1.64.finance-immersive..3.1.224....0.lQqDD8e_eWs#scso=uid_5XtPW6ifKeSkgge-6oc4_5:0,uid_9XtPW-7SC8PM_AaT5qawAg_5:0
#    pageStr="https://www.google.com/search?tbm=fin&ei=2ntPW7LGGM-p_QbjpqyYAQ&q="+ticker+"&oq="+ticker+"&gs_l=finance-immersive.3..81l3.9658.10169.0.10352.4.4.0.0.0.0.225.225.2-1.1.0..1..0...1.1.64.finance-immersive..3.1.224....0.lQqDD8e_eWs#scso=uid_5XtPW6ifKeSkgge-6oc4_5:0,uid_9XtPW-7SC8PM_AaT5qawAg_5:0"
#    page = requests.get(pageStr)
#    tree = html.fromstring(page.content)
#    return tree
    
#def GetMorningStar_MainPage(ticker):
#https://www.google.com/search?tbm=fin&ei=2ntPW7LGGM-p_QbjpqyYAQ&q=goog&oq=goog&gs_l=finance-immersive.3..81l3.9658.10169.0.10352.4.4.0.0.0.0.225.225.2-1.1.0..1..0...1.1.64.finance-immersive..3.1.224....0.lQqDD8e_eWs#scso=uid_5XtPW6ifKeSkgge-6oc4_5:0,uid_9XtPW-7SC8PM_AaT5qawAg_5:0
#    pageStr="https://www.google.com/search?tbm=fin&ei=2ntPW7LGGM-p_QbjpqyYAQ&q="+ticker+"&oq="+ticker+"&gs_l=finance-immersive.3..81l3.9658.10169.0.10352.4.4.0.0.0.0.225.225.2-1.1.0..1..0...1.1.64.finance-immersive..3.1.224....0.lQqDD8e_eWs#scso=uid_5XtPW6ifKeSkgge-6oc4_5:0,uid_9XtPW-7SC8PM_AaT5qawAg_5:0"
#    page = requests.get(pageStr)
#    tree = html.fromstring(page.content)
#    return tree
    
#def GetGoogleStock_Price(ticker):
    #tree = GetGoogle_MainPage(ticker)
    #price = round(float((tree.xpath('//*[@id="knowledge-finance-wholepage__entity-summary"]/div/g-card-section/div/g-card-section/div[1]/span[1]/span/span[1]')[0].text).replace(",","")),2)
    #return price

#def GetGoogleStock_PriceChange(ticker):
    #tree = GetGoogle_MainPage(ticker)
    #priceChange = tree.xpath('//*[@id="knowledge-finance-wholepage__entity-summary"]/div/g-card-section/div/g-card-section/div[1]/span[2]/span[1]')[0].text
    #return priceChange    
    
def GetBarChart_EarningsPage(ticker):
    #https://finance.yahoo.com/quote/AAPL?p=AAPL&.tsrc=fin-srch-v1
    pageStr="https://www.barchart.com/stocks/quotes/"+ticker+"/earnings-estimates"
    print 'pageStr=',pageStr
    #pageStr="https://www.barchart.com/stocks/quotes/EA/earnings-estimates"
    page = requests.get(pageStr)
    tree = html.fromstring(page.content)
    print tree
    return tree
    
def GetBarChart_Dates(tree):   
#    if(quarter_Pos==1):
#NEED CORRECT XPATH
#//*[@id="main-content-column"]/div/div[2]/div[2]/div[1]/div/div[2]/div/div/ng-transclude/table/tbody/tr[1]/td[2]
    #date = tree.xpath('/html[1]/body[1]/div[2]/div[1]/div[2]/div[2]/div[1]/div[2]/div[1]/div[1]/div[1]/div[2]/div[2]/div[1]/div[1]/div[2]/div[1]/div[1]/ng-transclude[1]/table[1]/tbody[1]/tr[1]/td[2]')
    #//*[@id="main-content-column"]/div/div[2]/div[2]/div[1]/div/div[2]/div/div/ng-transclude/table/thead/tr/th[2]/span[2]
    date = tree.xpath("//div[@class='earnings-table-content bc-table-wrapper']/barchart-table-scroll/table/thead/tr/th/span/text()")
    print 'date = ', date
    return date
    
def GetBarChart_Table(tree):   
    table = tree.xpath("//div[@class='earnings-table-content bc-table-wrapper']/barchart-table-scroll/table/tbody/tr/td/text()")
    print 'table2 = ', table
    return table    
    
def GetYahoo_MainPage(ticker):
    #https://finance.yahoo.com/quote/AAPL?p=AAPL&.tsrc=fin-srch-v1
    #Note: this string isnt exactly right, but it still works.   keep an eye on it
    pageStr="https://finance.yahoo.com/quote/"+ticker+"?p="+"ticker+&.tsrc=fin-srch-v1"
    #print pageStr
    page = requests.get(pageStr)
    tree = html.fromstring(page.content)
    return tree

def GetYahoo_StatisticsPage(ticker):
    #https://finance.yahoo.com/quote/GOOG/key-statistics?p=GOOG
    pageStr="https://finance.yahoo.com/quote/"+ticker+"/key-statistics?p="+ticker
    page = requests.get(pageStr)
    tree = html.fromstring(page.content)
    return tree    

def GetYahooPriceChangePct(tree)   :
    priceChange = tree.xpath('//*[@id="quote-header-info"]/div[3]/div[1]/div/span[2]')[0].text
    priceChangePct = priceChange[priceChange.find("(")+1:priceChange.find(")")]
    #print priceChangePct
    #gets rid of the percent
    priceChangePct = round(float(priceChangePct.replace("%", "")),3)
    return priceChangePct
    
def GetYahooLastClose_Price(tree):  
    #print 'tree = ',tree
    lastClosePrice = round(float((tree.xpath('//*[@id="quote-summary"]/div[1]/table/tbody/tr[1]/td[2]/span')[0].text).replace(",","")),2)
    #print lastClosePrice
    return lastClosePrice
    
def GetYahooStock_Price(tree):
    price = round(float((tree.xpath('//*[@id="quote-header-info"]/div[3]/div[1]/div/span[1]')[0].text).replace(",","")),2)
  #  print 'price : ', price
    return price
    
def GetYahooStock_FiftyDayMA(tree):
    fiftyDayMA = round(float((tree.xpath('//*[@id="Col1-0-KeyStatistics-Proxy"]/section/div[2]/div[2]/div/div[1]/table/tbody/tr[6]/td[2]')[0].text).replace(",","")),2)
    #print 'fiftyDayMA : ', fiftyDayMA
    return fiftyDayMA    

def GetYahooStock_TwoHundDayMA(tree):
    twohundDayMA = round(float((tree.xpath('//*[@id="Col1-0-KeyStatistics-Proxy"]/section/div[2]/div[2]/div/div[1]/table/tbody/tr[7]/td[2]')[0].text).replace(",","")),2)
    #print 'twohundDayMA : ', twohundDayMA
    return twohundDayMA        
  
    
def GetYahooStock_PriceChange(tree):
    #note: gives data in format of -.2717 (-.01%)  This funciton grabs would grab and return -.27 in this situation
    priceChange = tree.xpath('//*[@id="quote-header-info"]/div[3]/div[1]/div/span[2]')[0].text
    priceChange = round(float(priceChange.partition(' ')[0]),2)
    #print priceChange
   
    return priceChange   
    
def GetYahooStock_Beta(tree):
    beta = tree.xpath('//*[@id="quote-summary"]/div[2]/table/tbody/tr[2]/td[2]/span')[0].text
    #print 'beta = ', beta
    return beta

#Goes to the fidelity website, parses out the info we want and returns a dict
def StockParse():
    # Get Stock info
    page = requests.get('https://eresearch.fidelity.com/eresearch/gotoBL/fidelityTopOrders.jhtml', verify=False)
    tree = html.fromstring(page.content)

    #make arrays of length 30.  "1" is just initializing them all with the string "1".  This should not stay in the array.
    sTicker= ["1"] * 30
    sPriceChange = ["1"] * 30
    sBuy = ["1"] * 30
    sSell = ["1"] * 30
    sBuySellRatio = ["1"] * 30
    sCurPrice = ["1"] * 30
    sBeta = ["1"] * 30
    sPriceChangePercent = ["1"] *30
    sYahooPriceChange = ["1"] *30
    sLastClosePrice = ["1"] *30
    
    #Stock Lib is a dictionary.  It is how we return all of the values from this function.  Keep in mind it does not have labels.
    #Returns as #{{ticker1: price1, priceChange1, buy#1,sell#1,ratio#1},{ticker2: price2,priceChange2, buy#2,sell#2,ratio#2}, {3....}... }
    stockLib = {}

    #Start at i=1, end at 31.  (30 iterations)  Note: may give a syntax error in the morning b/c the website does not have 30 stocks listed
    for i in range (1,31):
        
        tickerListString = "//*[@id=\"topOrdersTable\"]/tbody/tr["
        tickerListString += str(i)
        tickerListString += "]/td[2]/span"

        priceChangeListString = "//*[@id=\"topOrdersTable\"]/tbody/tr["
        priceChangeListString += str(i)
        priceChangeListString += "]/td[4]/span"

        buyOrderListString = "//*[@id=\"topOrdersTable\"]/tbody/tr["
        buyOrderListString += str(i)
        buyOrderListString += "]/td[5]/span"
        
        sellOrderListString = "//*[@id=\"topOrdersTable\"]/tbody/tr["
        sellOrderListString += str(i)
        sellOrderListString += "]/td[7]/span"            

        #Actually get the fidelity data and put into arrays
        #i-1 because python arrays start at 0.
        sTicker[i-1] = tree.xpath(tickerListString)[0].text
        sPriceChange[i-1] = round(float((tree.xpath(priceChangeListString)[0].text).replace(",","")),2) #note, this is from fidelity and not updated always
        sBuy[i-1] = float(tree.xpath(buyOrderListString)[0].text)
        sSell[i-1] = float(tree.xpath(sellOrderListString)[0].text)
        
        #Yahoo Data
        treeYahoo = GetYahoo_MainPage(sTicker[i-1])
        
        sCurPrice[i-1] = float(GetYahooStock_Price(treeYahoo))
        sLastClosePrice[i-1] = float(GetYahooLastClose_Price(treeYahoo))
        sBeta[i-1] = GetYahooStock_Beta(treeYahoo)
        if(sBeta[i-1] != "N/A"):
            sBeta[i-1]=float(sBeta[i-1])
        sYahooPriceChange[i-1] = float(GetYahooStock_PriceChange(treeYahoo)) #Note: formatting is wrong.
        
        #calculated data
        #Round to 2nd decimal space
        sBuySellRatio[i-1] = round((float(sBuy[i-1]) / (float(sBuy[i-1]) + float(sSell[i-1]))),2)
        #sPriceChangePercent[i-1] = round(float(sYahooPriceChange[i-1] / sLastClosePrice[i-1])*100,2) #fidelity is not always updated.  Using Yahoo info instead
        sPriceChangePercent[i-1] = GetYahooPriceChangePct(treeYahoo)
        #print 'ticker = ', sTicker[i-1]
        #print 'price = ', sPriceChangePercent[i-1]
        
        #Add it to the stock Dictionary
        stockLib[sTicker[i-1]] = [sCurPrice[i-1], sYahooPriceChange[i-1], sPriceChangePercent[i-1], sBuy[i-1], sSell[i-1], sBuySellRatio[i-1], sBeta[i-1]]
    
    #Debug Prints
    #print 'sTicker ', sTicker
    #print 'sPriceChange ', sPriceChange
    #print 'sBuy ', sBuy
    #print 'sSell ', sSell
    #print 'sBuySellRatio ', sBuySellRatio        
    #print 'sYahooPriceChange ', sYahooPriceChange
    
    #print 'stockLib ', stockLib
    return stockLib
    
    
    #For XPARSE strings uncomment this!
    #print 'tickerListString: ' tickerListString
    #print 'PriceListString: ' priceListString
    #print 'buyOrderListString: ' buyOrderListString
    #print 'sellOrderListString: ' sellOrderListString

        #Line for stock company name
#        Stocks = tree.xpath('//*[@id="topOrdersTable"]/tbody/tr[3]/td[3]')[0].text
#        print 'Stocks: ', Stocks
#        print("Done")

def WriteTab_DailyStockList_old(workbook, stockLib, curTime):
    dateStr=str(curTime.month)+"_"+str(curTime.day)+"_"+str(curTime.year)     
    date_StockList_Title=dateStr+" Stock List"
    worksheet = workbook.add_worksheet(date_StockList_Title)

    green = workbook.add_format({'color': 'green'})
    red = workbook.add_format({'color': 'red'})
    
    worksheet.set_column('A:A', 10)
    worksheet.set_column('B:B', 10)
    worksheet.set_column('C:C', 15)
    worksheet.set_column('D:D', 15)
    worksheet.set_column('E:E', 12)
    worksheet.set_column('F:F', 10)
    worksheet.set_column('G:G', 15)
    worksheet.set_column('H:H', 10)
    
    worksheet.write('D1', 'Stock Additions')
    
    worksheet.write('A3', 'Ticker')             #Col 0
    worksheet.write('B3', 'Price')              #Col 1
    worksheet.write('C3', 'Price Change')       #Col 2
    worksheet.write('D3', 'Percent Change')     #Col 3
    worksheet.write('E3', 'Buy Orders')         #Col 4
    worksheet.write('F3', 'Sell Orders')        #Col 5
    worksheet.write('G3', 'Buy Sell Ratio')     #Col 6
    worksheet.write('H3', 'Beta')               #Col 7
    row=3
    for key in stockLib.keys():
        col=0
        worksheet.write(row,col,key)
        for item in stockLib[key]:               
            col=col+1
            #col 2 is Price Change, col 3 is price change pct, col 6 is buy sell ratio
            if (col == 2 or col == 3 or col == 6):
                if (item >= 0):
                    worksheet.write(row,col,item, green)
                else:
                    worksheet.write(row,col,item, red)
            else:
                worksheet.write(row,col,item)
        row=row+1

      
        
        
        
def WriteTab_DailyStockList(wb, stockLib, curTime):
    #sheet = wb.active
    sheets = wb.sheetnames
    print 'sheets0 : ' ,sheets
    sheet = wb[sheets[0]]
    dateStr=str(curTime.month)+"_"+str(curTime.day)+"_"+str(curTime.year)     
    date_StockList_Title=dateStr+" Stock List"
    
    green = Font(color=GREEN)
    red = Font(color=RED)
    #sheet.title = date_StockList_Title

    sheet['D1'] = "Stock Additions"         
    sheet['A3'] = "Ticker"                  #column 1
    sheet['B3'] = "Price"                   #column 2
    sheet['C3'] = "Price Change"            #column 3
    sheet['D3'] = "Percent Change"          #column 4
    sheet['E3'] = "Buy Orders"
    sheet['F3'] = "Sell Orders"
    sheet['G3'] = "Buy Sell Ratio"
    sheet['H3'] = "Beta"    
    
    #First Delete Rows (from yesterday)
    for row in sheet['A4:H35']:
        for cell in row:
            cell.value = None
            cell.font = None
            
    row_num=4
    for key in stockLib.keys():
        col_num=1
        sheet.cell(row=row_num, column=col_num).value = key
        for item in stockLib[key]:               
            col_num=col_num+1
            sheet.cell(row=row_num,column=col_num).value = item
            #col 2 is Price Change, col 3 is price change pct, col 6 is buy sell ratio
            if (col_num == 3 or col_num == 4):
                if (item >= 0):
                    sheet.cell(row=row_num,column=col_num).font = green
                else:
                    sheet.cell(row=row_num,column=col_num).font = red    
            if (col_num == 7):
                if (item >= .5):
                    sheet.cell(row=row_num,column=col_num).font = green
                else:
                    sheet.cell(row=row_num,column=col_num).font = red                    
        row_num=row_num+1
    
def WriteTab_CumulativeStockList(wb, stockLib, curTime):
#    sheet2 = wb.get_sheet_by_name("Cumulative_Stock_List")
    sheets = wb.sheetnames
    print 'sheets1: ' ,sheets
    sheet1 = wb[sheets[1]]
    dateStr=str(curTime.month)+"/"+str(curTime.day)+"/"+str(curTime.year)     
    
    green = Font(color=GREEN)
    red = Font(color=RED)
    
    #sheet1.title = "Cumulative_Stock_List"
    
    sheet1['A1'] = "Ticker"                  #column 2
    sheet1['B1'] = "Price"                   #column 3
    sheet1['C1'] = "Price Change"            #column 4
    sheet1['D1'] = "Percent Change"          #column 5
    sheet1['E1'] = "Buy Orders"
    sheet1['F1'] = "Sell Orders"
    sheet1['G1'] = "Buy Sell Ratio"
    sheet1['H1'] = "Beta" 
    sheet1['I1'] = "Date"         

    row_num=2
    lastRow=sheet1.max_row
    for key in stockLib.keys():
        col_num=1
        lastFoundMatch_Row=0
        for row_num in range(1, sheet1.max_row):
            #Look for last copy of ticker
            if key == sheet1.cell(row=row_num,column=col_num).value:
                #print 'found copy of ', key
                lastFoundMatch_Row=row_num            
            if(sheet1.cell(row=row_num,column=col_num).value == None):
                lastRow = row_num
                print 'lastRow = ', lastRow
                break
        #lastFoundMatch_Row is 0 if it is a new stock
        if lastFoundMatch_Row != 0:
            print 'Last match found on ', lastFoundMatch_Row
            sheet1.insert_rows(lastFoundMatch_Row+1)
            row_num=lastFoundMatch_Row+1
        else:
            row_num=lastRow
            print 'sheet1.maxrow +1', sheet1.max_row+1
        sheet1.cell(row=row_num, column=col_num).value = key
        #Insert the data (row_num gets updated if the stock is new or a copy)
        for item in stockLib[key]:               
            col_num=col_num+1
            sheet1.cell(row=row_num,column=col_num).value = item
            #col 2 is Price Change, col 3 is price change pct, col 6 is buy sell ratio
            if (col_num == 3 or col_num == 4):
                if (item >= 0):
                    sheet1.cell(row=row_num,column=col_num).font = green
                else:
                    sheet1.cell(row=row_num,column=col_num).font = red
            if ( col_num == 7):
                if (item >= .5):
                    sheet1.cell(row=row_num,column=col_num).font = green
                else:
                    sheet1.cell(row=row_num,column=col_num).font = red
        sheet1.cell(row=row_num,column=col_num+1).value = dateStr
        
def writeTemplate_Price_MAs(sheet2, stockLib, curTime):

    dateStr=str(curTime.month)+"/"+str(curTime.day)+"/"+str(curTime.year)     
    
    green = Font(color=GREEN)
    red = Font(color=RED) 

    #first input new date Also grab dates column
    dates_row=1
    for col_num in range (2,sheet2.max_column+5): #note 5 is an arbitrary number.  Needed incase we have NO dates and the max_column is 0
        if(sheet2.cell(row=dates_row,column=col_num).value == None):
            print 'putting date on row ', dates_row
            print 'putting date on col ', col_num
            #Fill in current Date
            sheet2.cell(row=dates_row, column=col_num).value = dateStr
            currentDate_Column=col_num
            break
            
    #see if our stock list has any new stocks
    stocks_col=1
    for key in stockLib.keys():
        matchFound=False        
        for row_num in range (2, sheet2.max_row+5):
            if(key == sheet2.cell(row=row_num, column=stocks_col).value):
                matchFound=True
            if(sheet2.cell(row=row_num, column=stocks_col).value == None):
                if (matchFound == False):
                    print 'did not find key: ',key
                    sheet2.cell(row=row_num, column=stocks_col).value = key
                else:
                    print '(SKIPPING)found key: ',key
                break
    return currentDate_Column
      
    
def WriteTabs_Price_MAs(wb, stockLib, curTime):
#These three tabs have the same template.  Should still compare the ticker list to make sure no errors occured
    #go by tab num
    green = Font(color=GREEN)
    red = Font(color=RED)
    
    price_ENUM=2
    fiftyDayMA_ENUM=3
    twohundDayMA_ENUM=4
    sheets = wb.sheetnames
    print 'sheets2: ' ,sheets
    sheet_Price = wb[sheets[price_ENUM]]
    sheet_fiftyDay = wb[sheets[fiftyDayMA_ENUM]]
    sheet_twohundDay = wb[sheets[twohundDayMA_ENUM]]
    currentDate_Column_price = writeTemplate_Price_MAs(sheet_Price, stockLib, curTime)
    currentDate_Column_fiftyDay = writeTemplate_Price_MAs(sheet_fiftyDay, stockLib, curTime)
    currentDate_Column_twohundDay = writeTemplate_Price_MAs(sheet_twohundDay, stockLib, curTime)

    #make sure dates are the same!
    if((currentDate_Column_price == currentDate_Column_fiftyDay) and (currentDate_Column_price == currentDate_Column_twohundDay)):
        currentDate_column = currentDate_Column_price
        print 'All column dates are the same.  Good! currentDate_column = ', currentDate_column
    else:
        print 'Column Dates are not the same, error!'
        print 'currentDate_Column_price = ', currentDate_Column_price
        print 'currentDate_Column_fiftyDay = ', currentDate_Column_fiftyDay
        print 'currentDate_Column_twohundDay = ', currentDate_Column_twohundDay
        
    cumulative_StockLib = {}
    
    stock_col=1
    stockMatch=True
    #make sure stock lists are the same!
    row_num=2
    while (sheet_Price.cell(row=row_num,column=stock_col).value != None):
        #print 'row_num= ', row_num
        if ((sheet_Price.cell(row=row_num,column=stock_col).value == sheet_fiftyDay.cell(row=row_num,column=stock_col).value) and  (sheet_Price.cell(row=row_num,column=stock_col).value == sheet_twohundDay.cell(row=row_num,column=stock_col).value)):
            #cumulativeStockLib{stockticker:price, 50dayma,200dayma}          
            
            stockTicker=str(sheet_Price.cell(row=row_num,column=stock_col).value)
            #Data from yahoo SATISTICS PAGE
            treeYahoo = GetYahoo_StatisticsPage(stockTicker)     
            print 'on stock : ', stockTicker

            #FIRST add in PRICE in form "190 (-.5%)"                
            stockPriceChange = float(GetYahooStock_PriceChange(treeYahoo))
            stockPrice = float(GetYahooStock_Price(treeYahoo))
            priceChangePercent = GetYahooPriceChangePct(treeYahoo)
            stockPriceTemplate = str(stockPrice) + " (" + str(priceChangePercent) + "%)"
            
            sheet_Price.cell(row=row_num,column=currentDate_column).value = stockPriceTemplate
            if (priceChangePercent >= 0):
                sheet_Price.cell(row=row_num,column=currentDate_column).font = green
            else:
                sheet_Price.cell(row=row_num,column=currentDate_column).font = red
                
            #Second Add in 50 Day MA
            fiftyDayMA = float(GetYahooStock_FiftyDayMA(treeYahoo))
            sheet_fiftyDay.cell(row=row_num,column=currentDate_column).value = fiftyDayMA
            
            #Add color if over /under moving averages
            if(fiftyDayMA > stockPrice):
                sheet_fiftyDay.cell(row=row_num,column=currentDate_column).font = green
            else:
                sheet_fiftyDay.cell(row=row_num,column=currentDate_column).font = red
                            
            #third Add in 200 Day MA
            twohundDayMA = float(GetYahooStock_TwoHundDayMA(treeYahoo))
            sheet_twohundDay.cell(row=row_num,column=currentDate_column).value = twohundDayMA

            #Add color if over /under moving averages
            if(twohundDayMA > stockPrice):
                sheet_twohundDay.cell(row=row_num,column=currentDate_column).font = green
            else:
                sheet_twohundDay.cell(row=row_num,column=currentDate_column).font = red            
                           
            #add to a dictionary.  Maybe will be useful? 
            cumulative_StockLib[stockTicker] = [priceChangePercent, stockPrice, stockPriceTemplate, fiftyDayMA, twohundDayMA]
        else:
            stockMatch=False
            break
        row_num=row_num+1
    if(stockMatch == True):
        print 'stocks match, continue'
    else:
        print 'ERROR: stocks Dont Match!'
    
    print 'cumulative_StockLibrary: ',cumulative_StockLib
    
    return cumulative_StockLib
    #check if stock exists
#    for key in stockLib.keys():

   
def WriteTab_Options(wb, stockLib, curTime):
    sheets = wb.sheetnames
    print 'sheets1: ' ,sheets
    sheet_options = wb[sheets[5]]
    
    #See if we already have options info on that stock, if not add it in
    stockCol_num=1
    for key in stockLib.keys():
        row_num=1
        found = 0
        while ((sheet_options.cell(row=row_num, column=stockCol_num).value != None) and (found  == 0)):
            if (key == sheet_options.cell(row=row_num, column=stockCol_num).value):
                found = 1
            else:
                found = 0
            row_num=row_num+16
        if (found == 1):
            print 'found key: ', key
        else:
            print 'could not find key: ', key
            print 'row_num = ', row_num
            newStockRow=row_num
            #insert stock info!
            sheet_options.cell(row=newStockRow, column=stockCol_num).value = key
            tree = GetBarChart_EarningsPage(key)
            earningsDates=[]
            earningsDates = GetBarChart_Dates(tree)
            if earningsDates: 
                earningsData=[]
                earningsData = GetBarChart_Table(tree)
            
                #First Table
                sheet_options.cell(row=newStockRow, column=stockCol_num + 1).value = 'Earnings History - Surprises'            
                sheet_options.cell(row=newStockRow + 2, column=stockCol_num + 1).value = 'Reported'
                sheet_options.cell(row=newStockRow + 3, column=stockCol_num + 1).value = 'Estimate'
                sheet_options.cell(row=newStockRow + 4, column=stockCol_num + 1).value = 'Difference'
                sheet_options.cell(row=newStockRow + 5, column=stockCol_num + 1).value = 'Surprise'
            
                #Earnings Dates Tbl 1
                print 'EARNINGDATES= ',earningsDates
                sheet_options.cell(row=newStockRow + 1, column=stockCol_num + 2).value = earningsDates[0] + " " + earningsDates[1]
                sheet_options.cell(row=newStockRow + 1, column=stockCol_num + 3).value = earningsDates[2] + " " + earningsDates[3]
                sheet_options.cell(row=newStockRow + 1, column=stockCol_num + 4).value = earningsDates[4] + " " + earningsDates[5]
                sheet_options.cell(row=newStockRow + 1, column=stockCol_num + 5).value = earningsDates[6] + " " + earningsDates[7]
                
                #Data Tbl 1
                sheet_options.cell(row=newStockRow + 2, column=stockCol_num + 2).value = earningsData[2]
                sheet_options.cell(row=newStockRow + 2, column=stockCol_num + 3).value = earningsData[3]
                sheet_options.cell(row=newStockRow + 2, column=stockCol_num + 4).value = earningsData[4]
                sheet_options.cell(row=newStockRow + 2, column=stockCol_num + 5).value = earningsData[5]
                
                sheet_options.cell(row=newStockRow + 3, column=stockCol_num + 2).value = earningsData[8]
                sheet_options.cell(row=newStockRow + 3, column=stockCol_num + 3).value = earningsData[9]
                sheet_options.cell(row=newStockRow + 3, column=stockCol_num + 4).value = earningsData[10]
                sheet_options.cell(row=newStockRow + 3, column=stockCol_num + 5).value = earningsData[11]
                
                sheet_options.cell(row=newStockRow + 4, column=stockCol_num + 2).value = earningsData[14]
                sheet_options.cell(row=newStockRow + 4, column=stockCol_num + 3).value = earningsData[15]
                sheet_options.cell(row=newStockRow + 4, column=stockCol_num + 4).value = earningsData[16]
                sheet_options.cell(row=newStockRow + 4, column=stockCol_num + 5).value = earningsData[17]  

                sheet_options.cell(row=newStockRow + 5, column=stockCol_num + 2).value = earningsData[20]
                sheet_options.cell(row=newStockRow + 5, column=stockCol_num + 3).value = earningsData[21]
                sheet_options.cell(row=newStockRow + 5, column=stockCol_num + 4).value = earningsData[22]
                sheet_options.cell(row=newStockRow + 5, column=stockCol_num + 5).value = earningsData[23]                
                
                #Second Table
                sheet_options.cell(row=newStockRow + 7, column=stockCol_num + 1).value = 'Earnings Estimates'            
                sheet_options.cell(row=newStockRow + 9, column=stockCol_num + 1).value = 'Average Estimate'
                sheet_options.cell(row=newStockRow + 10, column=stockCol_num + 1).value = 'Number of Estimates'
                sheet_options.cell(row=newStockRow + 11, column=stockCol_num + 1).value = 'High Estimate'
                sheet_options.cell(row=newStockRow + 12, column=stockCol_num + 1).value = 'Low Estimate'
                sheet_options.cell(row=newStockRow + 13, column=stockCol_num + 1).value = 'Prior Year'        
                sheet_options.cell(row=newStockRow + 14, column=stockCol_num + 1).value = 'Growth Rate Est. YoY'        

                #earnings dates tbl 2
                sheet_options.cell(row=newStockRow + 8, column=stockCol_num + 2).value = earningsDates[8] + " " + earningsDates[9]
                sheet_options.cell(row=newStockRow + 8, column=stockCol_num + 3).value = earningsDates[10] + " " + earningsDates[11]
                sheet_options.cell(row=newStockRow + 8, column=stockCol_num + 4).value = earningsDates[12] + " " + earningsDates[13]
                sheet_options.cell(row=newStockRow + 8, column=stockCol_num + 5).value = earningsDates[14] + " " + earningsDates[15]
                
                #data Tbl 2
                sheet_options.cell(row=newStockRow + 9, column=stockCol_num + 2).value = earningsData[26]
                sheet_options.cell(row=newStockRow + 9, column=stockCol_num + 3).value = earningsData[27]
                sheet_options.cell(row=newStockRow + 9, column=stockCol_num + 4).value = earningsData[28]
                sheet_options.cell(row=newStockRow + 9, column=stockCol_num + 5).value = earningsData[29]
                
                sheet_options.cell(row=newStockRow + 10, column=stockCol_num + 2).value = earningsData[32]
                sheet_options.cell(row=newStockRow + 10, column=stockCol_num + 3).value = earningsData[33]
                sheet_options.cell(row=newStockRow + 10, column=stockCol_num + 4).value = earningsData[34]
                sheet_options.cell(row=newStockRow + 10, column=stockCol_num + 5).value = earningsData[35]                
                
                sheet_options.cell(row=newStockRow + 11, column=stockCol_num + 2).value = earningsData[38]
                sheet_options.cell(row=newStockRow + 11, column=stockCol_num + 3).value = earningsData[39]
                sheet_options.cell(row=newStockRow + 11, column=stockCol_num + 4).value = earningsData[40]
                sheet_options.cell(row=newStockRow + 11, column=stockCol_num + 5).value = earningsData[41]               

                sheet_options.cell(row=newStockRow + 12, column=stockCol_num + 2).value = earningsData[44]
                sheet_options.cell(row=newStockRow + 12, column=stockCol_num + 3).value = earningsData[45]
                sheet_options.cell(row=newStockRow + 12, column=stockCol_num + 4).value = earningsData[46]
                sheet_options.cell(row=newStockRow + 12, column=stockCol_num + 5).value = earningsData[47]     

                sheet_options.cell(row=newStockRow + 13, column=stockCol_num + 2).value = earningsData[50]
                sheet_options.cell(row=newStockRow + 13, column=stockCol_num + 3).value = earningsData[51]
                sheet_options.cell(row=newStockRow + 13, column=stockCol_num + 4).value = earningsData[52]
                sheet_options.cell(row=newStockRow + 13, column=stockCol_num + 5).value = earningsData[53]        

                sheet_options.cell(row=newStockRow + 14, column=stockCol_num + 2).value = earningsData[57]
                sheet_options.cell(row=newStockRow + 14, column=stockCol_num + 3).value = earningsData[58]
                sheet_options.cell(row=newStockRow + 14, column=stockCol_num + 4).value = earningsData[59]
                sheet_options.cell(row=newStockRow + 14, column=stockCol_num + 5).value = earningsData[60]                     
            else:
                print 'stock not on list: ', key
                sheet_options.cell(row=newStockRow, column=stockCol_num + 1).value = 'STOCK NOT ON LIST!'            

            
            row_num=row_num+16

def WriteTotalChangeSinceInception(wb, cumulative_StockLib, curTime):
    sheets = wb.sheetnames
    #print 'sheets1: ' ,sheets
    sheet1 = wb[sheets[1]]
    dateStr=str(curTime.month)+"/"+str(curTime.day)+"/"+str(curTime.year)     
    
    green = Font(color=GREEN)
    red = Font(color=RED)
    
    col_num=1
    row_num = 1
    
    #get the last column to put the totals
    while (sheet1.cell(row=row_num,column=col_num).value != None and sheet1.cell(row=row_num,column=col_num).value != '% Change Since Inception'):
        col_num = col_num +1
    totals_col = col_num
    
    sheet1.cell(row=1,column=col_num).value = "% Change Since Inception"
    
    row_num = 1
    #column=1 is the tickers
    while (sheet1.cell(row=row_num,column=1).value != None):
        for key in cumulative_StockLib.keys():
            if (sheet1.cell(row=row_num,column=1).value == key):
                i=0
                for item in cumulative_StockLib[key]: 
                    if (i==0):
                        #print 'current price=', item
                        price_Change = item - sheet1.cell(row=row_num,column=2).value
                        pctChange = round(float(100*(price_Change/sheet1.cell(row=row_num,column=2).value)),2)
                        #print 'pctChange =', pctChange
                        sheet1.cell(row=row_num,column=totals_col).value = pctChange
                        if (pctChange > 0):
                            sheet1.cell(row=row_num,column=totals_col).font = green
                        else:
                            sheet1.cell(row=row_num,column=totals_col).font = red
                    i=i+1
        row_num=row_num+1
        
def WriteOverallStats(wb, cumulative_StockLib, curTime):
    sheets = wb.sheetnames
    #print 'sheets1: ' ,sheets
    sheet1 = wb[sheets[1]]
    dateStr=str(curTime.month)+"/"+str(curTime.day)+"/"+str(curTime.year)     
    
    green = Font(color=GREEN)
    red = Font(color=RED)
    
    col_num=1    
    while (sheet1.cell(row=1,column=col_num).value != '% Change Since Inception'):
        col_num = col_num +1
    totals_col = col_num
    
    #start at 2 because 1 is the '%change since inception'
    row_num = 2
    total_pctChange=0
    while (sheet1.cell(row=row_num,column=totals_col).value != None):
        total_pctChange=total_pctChange+sheet1.cell(row=row_num,column=totals_col).value
        row_num=row_num+1
        
    sheet1['P7'] = "Total Sum:"
    sheet1['Q7'] = str(total_pctChange) + "%"

def WriteFidelityStats(wb, cumulative_StockLib, curTime):
    sheets = wb.sheetnames
    #print 'sheets1: ' ,sheets
    sheet1 = wb[sheets[1]]
    dateStr=str(curTime.month)+"/"+str(curTime.day)+"/"+str(curTime.year)     
    
    green = Font(color=GREEN)
    red = Font(color=RED)

    col_num=1    
    while (sheet1.cell(row=1,column=col_num).value != 'Buy Sell Ratio'):
        col_num = col_num +1
    buysell_Ratio = col_num    
    
    col_num=1    
    while (sheet1.cell(row=1,column=col_num).value != '% Change Since Inception'):
        col_num = col_num +1
    totals_col = col_num
    
    row_num = 2
    greater_75_pctChange=0
    lower_30_pctChange=0
    btwn_35_65_pctChange=0
    while (sheet1.cell(row=row_num,column=totals_col).value != None):
        if(sheet1.cell(row=row_num,column=buysell_Ratio).value >= .75):
            print 'over .75 stock : ', sheet1.cell(row=row_num,column=1).value
            print 'totals col = ', totals_col
            print 'over .75 stock totals col = ', sheet1.cell(row=row_num,column=totals_col).value
            greater_75_pctChange=greater_75_pctChange+sheet1.cell(row=row_num,column=totals_col).value
        if (sheet1.cell(row=row_num,column=buysell_Ratio).value >= .35 and sheet1.cell(row=row_num,column=buysell_Ratio).value <= .65):
            btwn_35_65_pctChange=btwn_35_65_pctChange+sheet1.cell(row=row_num,column=totals_col).value
        if (sheet1.cell(row=row_num,column=buysell_Ratio).value <= .30):
            lower_30_pctChange=lower_30_pctChange+sheet1.cell(row=row_num,column=totals_col).value
        row_num=row_num+1
    
    sheet1['P8'] = ".75+ Ratio"
    sheet1['Q8'] = str(greater_75_pctChange) + "%"
    
    sheet1['P9'] = ".35-.65 Ratio"
    sheet1['Q9'] = str(btwn_35_65_pctChange) + "%"    
        
    sheet1['P10'] = ".30- Ratio"
    sheet1['Q10'] = str(lower_30_pctChange) + "%"
    
def csvWriter(stockLib):
    curTime = datetime.datetime.now()
    
    
    
    #workbook = xlsxwriter.Workbook('demo.xlsx')
    filepath='StockList.xlsx'
    #filepath='demo_2.xlsx'
    wb=load_workbook(filepath)
#    WriteTab_DailyStockList(workbook, stockLib, curTime)
    WriteTab_DailyStockList(wb, stockLib, curTime)
    
    WriteTab_CumulativeStockList(wb, stockLib, curTime)
    
    cumulative_StockLib = {}
    cumulative_StockLib = WriteTabs_Price_MAs(wb, stockLib, curTime)
        
    #NOTE: THIS SHOULD BE THE SAME AS stockLib, not cumulative_StockLib (Only for testing!)
    WriteTab_Options(wb, stockLib, curTime)

    WriteTotalChangeSinceInception(wb, cumulative_StockLib, curTime)
    
    WriteOverallStats(wb, cumulative_StockLib, curTime)

    WriteFidelityStats(wb, cumulative_StockLib, curTime)

    
    wb.save(filepath)
    #workbook.close()

    
   
    
def main():

    #Stock Lib is a dictionary.  It is how we return all of the values from this function.  Keep in mind it does not have labels.
    #Returns as #{{ticker1: price1, buy#1,sell#1,ratio#1},{ticker2: price2, buy#2,sell#2,ratio#2}, {3....}... }
    #Will need to figure out how to add yahoo functions to it as well.  Also need to figure out how to make it look nice in excel
    #earningDates=[]
    #tree = GetBarChart_EarningsPage('EA')
    #EaEarnings = []
    #EaEarnings = GetBarChart_Table(tree)
    #for i in range (0,len(EaEarnings)):
    #    print 'EaEarnings[i]',i, '=', EaEarnings[i]
    #print 'earning dates= ', earningDates

    #print 'earning dates[0] = ', earningDates[0]
    #print 'earning dates[1] = ', earningDates[1]
    #abc = earningDates[0] + " " + earningDates[1]
    #print 'abc = ', abc

    
    stockLib = {}
    stockLib = StockParse()
    
    for key in stockLib.keys():
        i=0
#        if key == 'ACBFF':
#            stockLib.pop(key,None)
#            print 'HARD-removing ACBFF'
        for item in stockLib[key]: 
            if (i==6):
                if (item > 5 and item != 'N/A'):
                    stockLib.pop(key,None)
                    print 'removing key: ', key
            i=i+1
    print 'stockLib in MAIN ', stockLib    
    
    csvWriter(stockLib)
    
    print 'Ticker: Price, PriceChange, price change %, buy amount, sell amount, ratio, beta'




if __name__ == '__main__':
    main()