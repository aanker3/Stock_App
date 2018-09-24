#!/usr/bin/env python3

# Name            :  StockParser
# Script Name     :  StockParser.py

import datetime
import re
from lxml import html
import requests
import time
import Tkinter
from PIL import ImageTk, Image
import os.path
import subprocess
import xlsxwriter #pip install xlsxwriter
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles.colors import RED
from openpyxl.styles.colors import GREEN
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter, column_index_from_string, coordinate_from_string
import copy
from openpyxl.worksheet import *
import sys

#google and morning star are not working for some reason.  List out of range.  has to do with the [0].text.  Not sure

#def GetGoogle_MainPage(ticker):
#https://www.google.com/search?tbm=fin&ei=2ntPW7LGGM-p_QbjpqyYAQ&q=goog&oq=goog&gs_l=finance-immersive.3..81l3.9658.10169.0.10352.4.4.0.0.0.0.225.225.2-1.1.0..1..0...1.1.64.finance-immersive..3.1.224....0.lQqDD8e_eWs#scso=uid_5XtPW6ifKeSkgge-6oc4_5:0,uid_9XtPW-7SC8PM_AaT5qawAg_5:0
#    pageStr="https://www.google.com/search?tbm=fin&ei=2ntPW7LGGM-p_QbjpqyYAQ&q="+ticker+"&oq="+ticker+"&gs_l=finance-immersive.3..81l3.9658.10169.0.10352.4.4.0.0.0.0.225.225.2-1.1.0..1..0...1.1.64.finance-immersive..3.1.224....0.lQqDD8e_eWs#scso=uid_5XtPW6ifKeSkgge-6oc4_5:0,uid_9XtPW-7SC8PM_AaT5qawAg_5:0"
#    page = requests.get(pageStr)
#    tree = html.fromstring(page.content)
#    return tree
    
#def GetMorningStar_MainPage(ticker):
#https://www.google.com/search?tbm=fin&ei=2ntPW7LGGM-p_QbjpqyYAQ&q=goog&oq=goog&gs_l=finance-immersive.3..81l3.9658.10169.0.10352.4.4.0.0.0.0.225.225.2-1.1.0..1..0...1.1.64.finance-immersive..3.1.224....0.lQqDD8e_eWs#scso=uid_5XtPW6ifKeSkgge-6oc4_5:0,uid_9XtPW-7SC8PM_AaT5qawAg_5:0
#    pageStr="https://www.google.com/search?tbm=fin&ei=2ntPW7LGGM-p_QbjpqyYAQ&q="+ticker+"&oq="+ticker+"&gs_l=finance-immersive.3..81l3.9658.10169.0.10352.4.4.0.0.0.0.225.225.2-1.1.0..1..0...1.1.64.finance-immersive..3.1.224....0.lQqDD8e_eWs#scso=uid_5XtPW6ifKeSkgge-6oc4_5:0,uid_9XtPW-7SC8PM_AaT5qawAg_5:0"
#    page = requests.get(pageStr)
#    tree = html.fromstring(page.content)
#    return tree
    
#def GetGoogleStock_Price(ticker):
    #tree = GetGoogle_MainPage(ticker)
    #price = round(float((tree.xpath('//*[@id="knowledge-finance-wholepage__entity-summary"]/div/g-card-section/div/g-card-section/div[1]/span[1]/span/span[1]')[0].text).replace(",","")),2)
    #return price

#def GetGoogleStock_PriceChange(ticker):
    #tree = GetGoogle_MainPage(ticker)
    #priceChange = tree.xpath('//*[@id="knowledge-finance-wholepage__entity-summary"]/div/g-card-section/div/g-card-section/div[1]/span[2]/span[1]')[0].text
    #return priceChange    
#stockLib_sBeta = 2
    
    
def GetBarChart_EarningsPage(ticker):
    #https://finance.yahoo.com/quote/AAPL?p=AAPL&.tsrc=fin-srch-v1
    pageStr="https://www.barchart.com/stocks/quotes/"+ticker+"/earnings-estimates"
    print 'pageStr=',pageStr
    #pageStr="https://www.barchart.com/stocks/quotes/EA/earnings-estimates"
    page = requests.get(pageStr)
    tree = html.fromstring(page.content)
    print tree
    return tree
    
def GetBarChart_Dates(tree):   
#    if(quarter_Pos==1):
#NEED CORRECT XPATH
#//*[@id="main-content-column"]/div/div[2]/div[2]/div[1]/div/div[2]/div/div/ng-transclude/table/tbody/tr[1]/td[2]
    #date = tree.xpath('/html[1]/body[1]/div[2]/div[1]/div[2]/div[2]/div[1]/div[2]/div[1]/div[1]/div[1]/div[2]/div[2]/div[1]/div[1]/div[2]/div[1]/div[1]/ng-transclude[1]/table[1]/tbody[1]/tr[1]/td[2]')
    #//*[@id="main-content-column"]/div/div[2]/div[2]/div[1]/div/div[2]/div/div/ng-transclude/table/thead/tr/th[2]/span[2]
    date = tree.xpath("//div[@class='earnings-table-content bc-table-wrapper']/barchart-table-scroll/table/thead/tr/th/span/text()")
    print 'date = ', date
    return date
    
def GetBarChart_Table(tree):   
    table = tree.xpath("//div[@class='earnings-table-content bc-table-wrapper']/barchart-table-scroll/table/tbody/tr/td/text()")
    print 'table2 = ', table
    return table    
    
def GetFinviz_StockPage(ticker):
    pageStr='https://finviz.com/quote.ashx?t='+ticker+'&ty=c&p=d&b=1'
    page = requests.get(pageStr)
    tree = html.fromstring(page.content)
    return tree
    
def GetFinvizStockINFO(tree):
    #gives in order of: ?perf week?, beta[0], ATR[1], previous close[2], price[3],
    stockInfo_1 = tree.xpath('//*[@class="table-dark-row"]/td[12]/b//text()') 
    print 'stockInfo_1 = ', stockInfo_1
    price=stockInfo_1[10]    
    beta=stockInfo_1[6]
    priceChangePct=str(stockInfo_1[11])
    print 'Before priceChangePct=',priceChangePct
    priceChangePct=priceChangePct.replace("%","")
    priceChangePct=priceChangePct.replace(" ","")


    print 'After priceChangePct=',priceChangePct
    
    

    #gives in order of: Shs Outstand [0], Shs Float[1], Short Float[2], Short Ratio[3], RSI(14)[4], Rel Volume[5], AVG VOlume[6], volume[7]
    stockInfo_3 = tree.xpath('//*[@class="table-dark-row"]/td[10]/b//text()')     
    sharesOutstand=stockInfo_3[0]
    rsi_14=stockInfo_3[8]
    avgVol=stockInfo_3[10]
    volume=stockInfo_3[11]
    targetPrice=stockInfo_3[4]
    high_52=stockInfo_3[6]
    low_52=stockInfo_3[7]

    
    #gives in order of: SMA200 (PCT)
    stockInfo_5 = tree.xpath('//*[@class="table-dark-row"]/td[8]/b//text()') 
    sma_200=stockInfo_5[11]
    profitMargin=stockInfo_5[9]
    
    #gives in order of: EPS (ttm)[0], EPS next Y[1], EPS next Q[2], EPS this Y[3], EPS Next Y [4], EPS next 5Y[5], Sales past 5Y[6], Sales Q/Q[7], EPS Q/Q[8], Earnings[9] 
    stockInfo_7 = tree.xpath('//*[@class="table-dark-row"]/td[6]/b//text()') 
    print 'stockInfo_7=',stockInfo_7
    eps_NextY=stockInfo_7[1]
    sales_QQ=stockInfo_7[8]
    eps_QQ=stockInfo_7[9]
    epsPast5Y=stockInfo_7[6]
    sma_50=stockInfo_7[11]
    
    #gives in order of:  P/B[0],P/C[1],quick Ratio[2], Current Ratio
    stockInfo_9 = tree.xpath('//*[@class="table-dark-row"]/td[4]/b//text()')     
    pe=stockInfo_9[0]
    forward_PE=stockInfo_9[1]
    ps=stockInfo_9[3]
    p_FCF=stockInfo_9[6]
    sma_20=stockInfo_9[11]

    #gives in order of: Index[0],Market Cap[1], Income[2], Sales[3], Book/sh[4], Cash/sh[5], Dividend[6], Dividend%[7], Employees[8], Optionable[9], Shortable[10], Recom[11]
    stockInfo_11 = tree.xpath('//*[@class="table-dark-row"]/td[2]/b/text()') 
    dividendPct=stockInfo_11[7] #BUG NOTE: This has been seen getting EMPLOYEES value
    if not (("%" in dividendPct) or ("-" in dividendPct)):
        dividendPct="known_bug"
    FinvizInfoDict={}
    FinvizInfoDict = {'Price': price, 'PriceChangePct':priceChangePct, 'beta': beta,'rsi_14':rsi_14, 'shares_Outstanding':sharesOutstand,  'avgVol':avgVol, 'volume':volume, 'targetPrice':targetPrice, 'high_52':high_52, 'low_52':low_52, 'sma_200':sma_200, 'eps_NextY':eps_NextY, 'sales_QQ':sales_QQ, 'eps_qq':eps_QQ, 'epsPast5Y':epsPast5Y, 'sma_50':sma_50, 'pe':pe, 'forward_PE':forward_PE,    'ps':ps, 'p_FCF':p_FCF, 'sma_20':sma_20, 'dividendPct':dividendPct, 'profitMargin':profitMargin}
    print FinvizInfoDict
    return FinvizInfoDict

def GetFinvizStockINFO_test(tree):
    #gives in order of: 
    print tree.xpath('//*[@class="table-dark-row"]/td[12]/b//text()')

    
    #print stockInfo
    #return priceChangePct     
    
def GetYahoo_MainPage(ticker):
    #https://finance.yahoo.com/quote/AAPL?p=AAPL&.tsrc=fin-srch-v1
    #Note: this string isnt exactly right, but it still works.   keep an eye on it
    pageStr="https://finance.yahoo.com/quote/"+ticker+"?p="+"ticker+&.tsrc=fin-srch-v1"
    #print pageStr
    page = requests.get(pageStr)
    tree = html.fromstring(page.content)
    return tree

def GetYahoo_StatisticsPage(ticker):
    #https://finance.yahoo.com/quote/GOOG/key-statistics?p=GOOG
    pageStr="https://finance.yahoo.com/quote/"+ticker+"/key-statistics?p="+ticker
    page = requests.get(pageStr)
    tree = html.fromstring(page.content)
    return tree    

def GetYahooPriceChangePct(tree)   :
    priceChange = tree.xpath('//*[@id="quote-header-info"]/div[3]/div[1]/div/span[2]')[0].text
    priceChangePct = priceChange[priceChange.find("(")+1:priceChange.find(")")]
    #print priceChangePct
    #gets rid of the percent
    priceChangePct = round(float(priceChangePct.replace("%", "")),3)
    return priceChangePct
    
def GetYahooLastClose_Price(tree):  
    #print 'tree = ',tree
    lastClosePrice = round(float((tree.xpath('//*[@id="quote-summary"]/div[1]/table/tbody/tr[1]/td[2]/span')[0].text).replace(",","")),2)
    #print lastClosePrice
    return lastClosePrice
    
def GetYahooStock_Price(tree):
    price = round(float((tree.xpath('//*[@id="quote-header-info"]/div[3]/div[1]/div/span[1]')[0].text).replace(",","")),2)
  #  print 'price : ', price
    return price
    
def GetYahooStock_FiftyDayMA(tree):
    fiftyDayMA = round(float((tree.xpath('//*[@id="Col1-0-KeyStatistics-Proxy"]/section/div[2]/div[2]/div/div[1]/table/tbody/tr[6]/td[2]')[0].text).replace(",","")),2)
    #print 'fiftyDayMA : ', fiftyDayMA
    return fiftyDayMA    

def GetYahooStock_TwoHundDayMA(tree):
    twohundDayMA = round(float((tree.xpath('//*[@id="Col1-0-KeyStatistics-Proxy"]/section/div[2]/div[2]/div/div[1]/table/tbody/tr[7]/td[2]')[0].text).replace(",","")),2)
    #print 'twohundDayMA : ', twohundDayMA
    return twohundDayMA        
  
    
def GetYahooStock_PriceChange(tree):
    #note: gives data in format of -.2717 (-.01%)  This funciton grabs would grab and return -.27 in this situation
    priceChange = tree.xpath('//*[@id="quote-header-info"]/div[3]/div[1]/div/span[2]')[0].text
    priceChange = round(float(priceChange.partition(' ')[0]),2)
    #print priceChange
   
    return priceChange   
    
def GetYahooStock_Beta(tree):
    beta = tree.xpath('//*[@id="quote-summary"]/div[2]/table/tbody/tr[2]/td[2]/span')[0].text
    #print 'beta = ', beta
    return beta

def FinvizStockParse(web):    
    # Get Stock info
    page = requests.get(web, verify=False)
    tree = html.fromstring(page.content)


    #ticker = tree.xpath('//*[@class="table-dark-row-cp"]/td/a/text()') #(gets every other all stuff)
    tickers_Even = tree.xpath('//*[@class="table-dark-row-cp"]/td[1]/a/text()') #gets odds
    #print 'tickers_Even=',tickers_Even
    tickers_Odd = tree.xpath('//*[@class="table-light-row-cp"]/td[1]/a/text()') #gets evens
    sectors_Even = tree.xpath('//*[@class="table-dark-row-cp"]/td[3]/a/text()') #gets odds
    #print 'sectors_Even = ', sectors_Even
    sectors_Odd = tree.xpath('//*[@class="table-light-row-cp"]/td[3]/a/text()') #gets evens
    industries_Even = tree.xpath('//*[@class="table-dark-row-cp"]/td[4]/a/text()') #gets odds
    industries_Odd = tree.xpath('//*[@class="table-light-row-cp"]/td[4]/a/text()') #gets evens
    prices_Odd = tree.xpath('//*[@class="table-light-row-cp"]/td[8]//text()') #gets ODDS
    #print 'prices_Odd=', prices_Odd
    prices_Even = tree.xpath('//*[@class="table-dark-row-cp"]/td[8]//text()') #gets evens
    priceChangePct_Odd = tree.xpath('//*[@class="table-light-row-cp"]/td[9]//text()') #gets ODDS
    #print 'priceChangePct_Odd=', priceChangePct_Odd
    priceChangePct_Even = tree.xpath('//*[@class="table-dark-row-cp"]/td[9]//text()') #get
    #beta_Odd = tree.xpath('//*[@class="table-light-row-cp"]/td[6]//text()') #gets 
    rsi_Odd = tree.xpath('//*[@class="table-light-row-cp"]/td[7]//text()') #gets ODDS
    #print 'rsi_odd=', rsi_Odd
    rsi_Even = tree.xpath('//*[@class="table-dark-row-cp"]/td[7]//text()') #gets EVENS
    #print 'sectors_Even = ', sectors_Even

    stockLib = {}
    #make arrays of length 29.  "1" is just initializing them all with the string "1".  This should not stay in the array.
    sTicker= ["1"] * 20
    sCurPrice = ["1"] * 20
    sRSI = ["1"] * 20
    sPriceChangePercent = ["1"] *20
    sYahooPriceChange = ["1"] *20
    sSector = ["1"]*20
    sIndustry = ["1"]*20
    
    #BUG: IF Finviz stocks are not full
    
    #do it twice for evens and odds
    for i in range (0,len(tickers_Even)):
        sTicker[i]=tickers_Even[i]
        sSector[i]=sectors_Even[i]
        sIndustry[i]=industries_Even[i]
        sCurPrice[i]=float(prices_Even[i])
        sPriceChangePercent[i]=float(priceChangePct_Even[i].replace("%",""))
        sRSI[i]=rsi_Even[i] 
        if (sRSI[i] != "-"):
            sRSI[i] = float(sRSI[i])        
    for i in range (len(tickers_Even),len(tickers_Even)+len(tickers_Odd)):
        sTicker[i]=tickers_Odd[i-len(tickers_Even)]
        sSector[i]=sectors_Odd[i-len(tickers_Even)]
        sIndustry[i]=industries_Odd[i-len(tickers_Even)]
        sCurPrice[i]=float(prices_Odd[i-len(tickers_Even)])   
        sPriceChangePercent[i]=float(priceChangePct_Odd[i-len(tickers_Even)].replace("%",""))     
        sRSI[i]=rsi_Odd[i-len(tickers_Even)]    
        if (sRSI[i] != "-"):
            sRSI[i] = float(sRSI[i])
    #print 'sTicker =', sTicker
    
    #get all other data
    for i in range (0,len(tickers_Even) + len(tickers_Odd)):
        #Yahoo Data
        #treeYahoo = GetYahoo_MainPage(sTicker[i])
        #sCurPrice[i] = float(GetYahooStock_Price(treeYahoo))
        #sYahooPriceChange[i] = float(GetYahooStock_PriceChange(treeYahoo))
        #sPriceChangePercent[i] = GetYahooPriceChangePct(treeYahoo)
        #sBeta[i] = GetYahooStock_Beta(treeYahoo)
        #if(sBeta[i] != "N/A"):
        #    sBeta[i]=float(sBeta[i])        
        #Stock Lib is a dictionary.  It is how we return all of the values from this function.  Keep in mind it does not have labels.
        #Returns as #{{ticker1: price1, priceChange1, priceChangePct1, beta1},{ticker2: price2,priceChange2, pricechangepct#2,beta#2}, {3....}... }
                    
        stockLib[sTicker[i]] = [sCurPrice[i], sPriceChangePercent[i], sRSI[i], sSector[i], sIndustry[i]]
    #   print stockLib
    return stockLib

            
            
#Goes to the fidelity website, parses out the info we want and returns a dict
def StockParse():
    # Get Stock info
    page = requests.get('https://eresearch.fidelity.com/eresearch/gotoBL/fidelityTopOrders.jhtml', verify=False)
    tree = html.fromstring(page.content)
    
    #make arrays of length 30.  "1" is just initializing them all with the string "1".  This should not stay in the array.
    sTicker= ["1"] * 30
    sPriceChange = ["1"] * 30
    sBuy = ["1"] * 30
    sSell = ["1"] * 30
    sBuySellRatio = ["1"] * 30
    sCurPrice = ["1"] * 30
    sBeta = ["1"] * 30
    sPriceChangePercent = ["1"] *30
    sYahooPriceChange = ["1"] *30
    sLastClosePrice = ["1"] *30
    
    #Stock Lib is a dictionary.  It is how we return all of the values from this function.  Keep in mind it does not have labels.
    #Returns as #{{ticker1: price1, priceChange1, buy#1,sell#1,ratio#1},{ticker2: price2,priceChange2, buy#2,sell#2,ratio#2}, {3....}... }
    stockLib = {}

    #Start at i=1, end at 31.  (30 iterations)  Note: may give a syntax error in the morning b/c the website does not have 30 stocks listed
    for i in range (1,31):
        
        tickerListString = "//*[@id=\"topOrdersTable\"]/tbody/tr["
        tickerListString += str(i)
        tickerListString += "]/td[2]/span"

        priceChangeListString = "//*[@id=\"topOrdersTable\"]/tbody/tr["
        priceChangeListString += str(i)
        priceChangeListString += "]/td[4]/span"

        buyOrderListString = "//*[@id=\"topOrdersTable\"]/tbody/tr["
        buyOrderListString += str(i)
        buyOrderListString += "]/td[5]/span"
        
        sellOrderListString = "//*[@id=\"topOrdersTable\"]/tbody/tr["
        sellOrderListString += str(i)
        sellOrderListString += "]/td[7]/span"            

        #Actually get the fidelity data and put into arrays
        #i-1 because python arrays start at 0.
        sTicker[i-1] = tree.xpath(tickerListString)[0].text
        sPriceChange[i-1] = round(float((tree.xpath(priceChangeListString)[0].text).replace(",","")),2) #note, this is from fidelity and not updated always
        sBuy[i-1] = float(tree.xpath(buyOrderListString)[0].text)
        sSell[i-1] = float(tree.xpath(sellOrderListString)[0].text)
        
        #Yahoo Data
        treeYahoo = GetYahoo_MainPage(sTicker[i-1])
        
        sCurPrice[i-1] = float(GetYahooStock_Price(treeYahoo))
        sLastClosePrice[i-1] = float(GetYahooLastClose_Price(treeYahoo))
        sBeta[i-1] = GetYahooStock_Beta(treeYahoo)
        if(sBeta[i-1] != "N/A"):
            sBeta[i-1]=float(sBeta[i-1])
        sYahooPriceChange[i-1] = float(GetYahooStock_PriceChange(treeYahoo)) #Note: formatting is wrong.
        
        #calculated data
        #Round to 2nd decimal space
        sBuySellRatio[i-1] = round((float(sBuy[i-1]) / (float(sBuy[i-1]) + float(sSell[i-1]))),2)
        #sPriceChangePercent[i-1] = round(float(sYahooPriceChange[i-1] / sLastClosePrice[i-1])*100,2) #fidelity is not always updated.  Using Yahoo info instead
        sPriceChangePercent[i-1] = GetYahooPriceChangePct(treeYahoo)
        #print 'ticker = ', sTicker[i-1]
        #print 'price = ', sPriceChangePercent[i-1]
        
        #Add it to the stock Dictionary
        stockLib[sTicker[i-1]] = [sCurPrice[i-1], sYahooPriceChange[i-1], sPriceChangePercent[i-1], sBuy[i-1], sSell[i-1], sBuySellRatio[i-1], sBeta[i-1]]
    
    #Debug Prints
    #print 'sTicker ', sTicker
    #print 'sPriceChange ', sPriceChange
    #print 'sBuy ', sBuy
    #print 'sSell ', sSell
    #print 'sBuySellRatio ', sBuySellRatio        
    #print 'sYahooPriceChange ', sYahooPriceChange
    
    #print 'stockLib ', stockLib
    return stockLib
    
    
    #For XPARSE strings uncomment this!
    #print 'tickerListString: ' tickerListString
    #print 'PriceListString: ' priceListString
    #print 'buyOrderListString: ' buyOrderListString
    #print 'sellOrderListString: ' sellOrderListString

        #Line for stock company name
#        Stocks = tree.xpath('//*[@id="topOrdersTable"]/tbody/tr[3]/td[3]')[0].text
#        print 'Stocks: ', Stocks
#        print("Done")

def WriteTab_DailyStockList_old(workbook, stockLib, curTime):
    dateStr=str(curTime.month)+"_"+str(curTime.day)+"_"+str(curTime.year)     
    date_StockList_Title=dateStr+" Stock List"
    worksheet = workbook.add_worksheet(date_StockList_Title)

    green = workbook.add_format({'color': 'green'})
    red = workbook.add_format({'color': 'red'})
    
    worksheet.set_column('A:A', 10)
    worksheet.set_column('B:B', 10)
    worksheet.set_column('C:C', 15)
    worksheet.set_column('D:D', 15)
    worksheet.set_column('E:E', 12)
    worksheet.set_column('F:F', 10)
    worksheet.set_column('G:G', 15)
    worksheet.set_column('H:H', 10)
    
    worksheet.write('D1', 'Stock Additions')
    
    worksheet.write('A3', 'Ticker')             #Col 0
    worksheet.write('B3', 'Price')              #Col 1
    worksheet.write('C3', 'Price Change')       #Col 2
    worksheet.write('D3', 'Percent Change')     #Col 3
    worksheet.write('E3', 'Buy Orders')         #Col 4
    worksheet.write('F3', 'Sell Orders')        #Col 5
    worksheet.write('G3', 'Buy Sell Ratio')     #Col 6
    worksheet.write('H3', 'Beta')               #Col 7
    row=3
    for key in stockLib.keys():
        col=0
        worksheet.write(row,col,key)
        for item in stockLib[key]:               
            col=col+1
            #col 2 is Price Change, col 3 is price change pct, col 6 is buy sell ratio
            if (col == 2 or col == 3 or col == 6):
                if (item >= 0):
                    worksheet.write(row,col,item, green)
                else:
                    worksheet.write(row,col,item, red)
            else:
                worksheet.write(row,col,item)
        row=row+1

      
        
        
        
def WriteTab_DailyStockList(wb, stockLib, curTime):
    #sheet = wb.active
    sheets = wb.sheetnames
    print 'DAILYSTOCKLIST sheets0 : ' ,sheets
    sheet = wb[sheets[0]]
    dateStr=str(curTime.month)+"_"+str(curTime.day)+"_"+str(curTime.year)     
    date_StockList_Title=dateStr+" Stock List"
    
    green = Font(color=GREEN)
    red = Font(color=RED)
    #sheet.title = date_StockList_Title

    sheet['D1'] = "Stock Additions"         
    sheet['A3'] = "Ticker"                  #column 1
    sheet['B3'] = "Price"                   #column 2
    sheet['C3'] = "Percent Change"          #column 4
    sheet['D3'] = "Rsi"    
    sheet['E3'] = "Sector"
    sheet['F3'] = "Industry"    
    
    ticker_Col=1
    price_Col=2
    pctChange_Col=3
    Rsi_Col=4
    Sector_Col=5
    Industry_Col=6
    
    #First Delete Rows (from yesterday)
    for row in sheet['A4:H35']:
        for cell in row:
            cell.value = None
            cell.font = None
            
    row_num=4
    for key in stockLib.keys():
        print 'daily stocklist on key: ', key
        col_num=1
        sheet.cell(row=row_num, column=col_num).value = key
        for item in stockLib[key]:               
            col_num=col_num+1
            sheet.cell(row=row_num,column=col_num).value = item
            #col 2 is Price Change, col 3 is price change pct, col 6 is buy sell ratio
            if (col_num == pctChange_Col):
                if (item >= 0):
                    sheet.cell(row=row_num,column=col_num).font = green
                else:
                    sheet.cell(row=row_num,column=col_num).font = red        
            if (col_num == Rsi_Col):
                if (item <= 30):
                    sheet.cell(row=row_num,column=col_num).font = green
                elif (item >=70):
                    sheet.cell(row=row_num,column=col_num).font = red    
                else:
                    sheet.cell(row=row_num,column=col_num).font = None                      
        row_num=row_num+1
    
def WriteTab_CumulativeStockList(wb, stockLib, curTime):
#    sheet2 = wb.get_sheet_by_name("Cumulative_Stock_List")
    sheets = wb.sheetnames
    print 'sheets1: ' ,sheets
    sheet1 = wb[sheets[1]]
    dateStr=str(curTime.month)+"/"+str(curTime.day)+"/"+str(curTime.year)     
    
    green = Font(color=GREEN)
    red = Font(color=RED)
    
    #sheet1.title = "Cumulative_Stock_List"

    ticker_Col=1
    price_Col=2
    pctChange_Col=3
    Rsi_Col=4
    Sector_Col=5
    Industry_Col=6
    Date_Col=7
        
    sheet1.cell(row=1,column=ticker_Col).value = "Ticker"                  #column 2
    sheet1.cell(row=1,column=price_Col).value = "Price"                   #column 3
    sheet1.cell(row=1,column=pctChange_Col).value = "Percent Change"          #column 5
    sheet1.cell(row=1,column=Rsi_Col).value  = "Rsi" 
    sheet1.cell(row=1,column=Sector_Col).value  = "Sector"
    sheet1.cell(row=1,column=Industry_Col).value  = "Industry"
    sheet1.cell(row=1,column=Date_Col).value  = "Date" 


    row_num=2
    lastRow=sheet1.max_row
    for key in stockLib.keys():
        col_num=1
        lastFoundMatch_Row=0
        row_num=1
        while(sheet1.cell(row=row_num,column=col_num).value != None):
            #Look for last copy of ticker
            if key == sheet1.cell(row=row_num,column=col_num).value:
                #print 'found copy of ', key
                lastFoundMatch_Row=row_num            
            row_num=row_num+1
        if(sheet1.cell(row=row_num,column=col_num).value == None):
            lastRow = row_num
            print 'lastRow = ', lastRow
        #lastFoundMatch_Row is 0 if it is a new stock
        if lastFoundMatch_Row != 0:
            print 'Last match found on ', lastFoundMatch_Row
            sheet1.insert_rows(lastFoundMatch_Row+1)
            row_num=lastFoundMatch_Row+1
        else:
            row_num=lastRow
            print 'sheet1.maxrow +1', sheet1.max_row+1
        sheet1.cell(row=row_num, column=col_num).value = key
        #Insert the data (row_num gets updated if the stock is new or a copy)
        for item in stockLib[key]:               
            col_num=col_num+1
            sheet1.cell(row=row_num,column=col_num).value = item
            #col 2 is Price Change, col 3 is price change pct, col 6 is buy sell ratio
            if (col_num == pctChange_Col):
                if (float(item) >= 0):
                    sheet1.cell(row=row_num,column=col_num).font = green
                else:
                    sheet1.cell(row=row_num,column=col_num).font = red                    
            if (col_num == Rsi_Col):
                if (item <= 30):
                    sheet1.cell(row=row_num,column=col_num).font = green
                elif (item >=70):
                    sheet1.cell(row=row_num,column=col_num).font = red    
                else:
                    sheet1.cell(row=row_num,column=col_num).font = None  
        sheet1.cell(row=row_num,column=col_num+1).value = dateStr
        
def writeTemplate_Price_MAs(sheet2, stockLib, curTime):

    dateStr=str(curTime.month)+"/"+str(curTime.day)+"/"+str(curTime.year)     
    sheet2.cell(row=1,column=1).value = "Stock/Date"
    green = Font(color=GREEN)
    red = Font(color=RED) 

    #first input new date Also grab dates column
    dates_row=1
    for col_num in range (2,sheet2.max_column+5): #note 5 is an arbitrary number.  Needed incase we have NO dates and the max_column is 0
        if(sheet2.cell(row=dates_row,column=col_num).value == None):
            print 'putting date on row ', dates_row
            print 'putting date on col ', col_num
            #Fill in current Date
            sheet2.cell(row=dates_row, column=col_num).value = dateStr
            currentDate_Column=col_num
            break
            
    #see if our stock list has any new stocks
    stocks_col=1
    for key in stockLib.keys():
        matchFound=False        
        for row_num in range (2, sheet2.max_row+5):
            if(key == sheet2.cell(row=row_num, column=stocks_col).value):
                matchFound=True
            if(sheet2.cell(row=row_num, column=stocks_col).value == None):
                if (matchFound == False):
                    print 'did not find key: ',key
                    sheet2.cell(row=row_num, column=stocks_col).value = key
                else:
                    print '(SKIPPING)found key: ',key
                break
    return currentDate_Column
      
def FinvizListSite(sheet_Price, stockticker, listOfStockDicts):
    print 'in FinvizListSite, stockticker=', stockticker
    FinvizListAddr='https://finviz.com/screener.ashx?v=152&t='
    for i in range(0,len(stockticker)):
        FinvizListAddr=FinvizListAddr+stockticker[i]
        if (i+1 != len(stockticker)):
            FinvizListAddr=FinvizListAddr+','
    FinvizListAddr=FinvizListAddr+'&c=1,7,8,10,13,14,18,19,22,23,24,41,48,52,53,54,57,58,59,63,65,66,67,69'
    
    print 'FinvizListAddr=',FinvizListAddr

    page = requests.get(FinvizListAddr)
    tree = html.fromstring(page.content)        

    tickers_Even = tree.xpath('//*[@class="table-dark-row-cp"]/td[1]/a/text()') #gets odds
    tickers_Odd = tree.xpath('//*[@class="table-light-row-cp"]/td[1]/a/text()') #gets odds
    pe_Even = tree.xpath('//*[@class="table-dark-row-cp"]/td[2]//text()') #gets odds
    pe_Odd = tree.xpath('//*[@class="table-light-row-cp"]/td[2]//text()') #gets odds    
    forwardPE_Even = tree.xpath('//*[@class="table-dark-row-cp"]/td[3]//text()') #gets odds
    forwardPE_Odd = tree.xpath('//*[@class="table-light-row-cp"]/td[3]//text()') #gets odds
    ps_Even = tree.xpath('//*[@class="table-dark-row-cp"]/td[4]//text()') #gets odds
    ps_Odd = tree.xpath('//*[@class="table-light-row-cp"]/td[4]//text()') #gets odds
    pFCF_Even = tree.xpath('//*[@class="table-dark-row-cp"]/td[5]//text()') #gets odds
    pFCF_Odd = tree.xpath('//*[@class="table-light-row-cp"]/td[5]//text()') #gets odds
    dividendPct_Even= tree.xpath('//*[@class="table-dark-row-cp"]/td[6]//text()') #gets odds
    dividendPct_Odd = tree.xpath('//*[@class="table-light-row-cp"]/td[6]//text()') #gets odds
    epsNextY_Even = tree.xpath('//*[@class="table-dark-row-cp"]/td[7]//text()') #gets odds
    epsNextY_Odd = tree.xpath('//*[@class="table-light-row-cp"]/td[7]//text()') #gets odds
    epsPast5Y_Even =tree.xpath('//*[@class="table-dark-row-cp"]/td[8]//text()') #gets odds
    epsPast5Y_Odd  = tree.xpath('//*[@class="table-light-row-cp"]/td[8]//text()') #gets odds
    epsQQ_Even = tree.xpath('//*[@class="table-dark-row-cp"]/td[9]//text()') #gets odds
    epsQQ_Odd = tree.xpath('//*[@class="table-light-row-cp"]/td[9]//text()') #gets odds
    salesQQ_Even= tree.xpath('//*[@class="table-dark-row-cp"]/td[10]//text()') #gets odds
    salesQQ_Odd = tree.xpath('//*[@class="table-light-row-cp"]/td[10]//text()') #gets odds
    sharesOutstanding_Even= tree.xpath('//*[@class="table-dark-row-cp"]/td[11]//text()') #gets odds
    sharesOutstanding_Odd = tree.xpath('//*[@class="table-light-row-cp"]/td[11]//text()') #gets odds
    profitMargin_Even= tree.xpath('//*[@class="table-dark-row-cp"]/td[12]//text()') #gets odds
    profitMargin_Odd = tree.xpath('//*[@class="table-light-row-cp"]/td[12]//text()') #gets odds    
    beta_Even= tree.xpath('//*[@class="table-dark-row-cp"]/td[13]//text()') #gets odds
    beta_Odd = tree.xpath('//*[@class="table-light-row-cp"]/td[13]//text()') #gets odds       
    sma20_Even= tree.xpath('//*[@class="table-dark-row-cp"]/td[14]//text()') #gets odds
    sma20_Odd = tree.xpath('//*[@class="table-light-row-cp"]/td[14]//text()') #gets odds   
    sma50_Even= tree.xpath('//*[@class="table-dark-row-cp"]/td[15]//text()') #gets odds
    sma50_Odd = tree.xpath('//*[@class="table-light-row-cp"]/td[15]//text()') #gets odds   
    sma200_Even= tree.xpath('//*[@class="table-dark-row-cp"]/td[16]//text()') #gets odds
    sma200_Odd = tree.xpath('//*[@class="table-light-row-cp"]/td[16]//text()') #gets odds   
    high52W_Even =tree.xpath('//*[@class="table-dark-row-cp"]/td[17]//text()') #gets odds
    high52W_Odd = tree.xpath('//*[@class="table-light-row-cp"]/td[17]//text()') #gets odds   
    low52W_Even =tree.xpath('//*[@class="table-dark-row-cp"]/td[18]//text()') #gets odds
    low52W_Odd = tree.xpath('//*[@class="table-light-row-cp"]/td[18]//text()') #gets odds   
    rsi_Even =tree.xpath('//*[@class="table-dark-row-cp"]/td[19]//text()') #gets odds
    rsi_Odd = tree.xpath('//*[@class="table-light-row-cp"]/td[19]//text()') #gets odds     
    avgVol_Even =tree.xpath('//*[@class="table-dark-row-cp"]/td[20]//text()') #gets odds
    avgVol_Odd = tree.xpath('//*[@class="table-light-row-cp"]/td[20]//text()') #gets odds         
    price_Even =tree.xpath('//*[@class="table-dark-row-cp"]/td[21]//text()') #gets odds
    price_Odd = tree.xpath('//*[@class="table-light-row-cp"]/td[21]//text()') #gets odds     
    priceChangePct_Even =tree.xpath('//*[@class="table-dark-row-cp"]/td[22]//text()') #gets odds
    priceChangePct_Odd = tree.xpath('//*[@class="table-light-row-cp"]/td[22]//text()') #gets odds       
    volume_Even =tree.xpath('//*[@class="table-dark-row-cp"]/td[23]//text()') #gets odds
    volume_Odd = tree.xpath('//*[@class="table-light-row-cp"]/td[23]//text()') #gets odds     
    targetPrice_Even =tree.xpath('//*[@class="table-dark-row-cp"]/td[24]//text()') #gets odds
    targetPrice_Odd = tree.xpath('//*[@class="table-light-row-cp"]/td[24]//text()') #gets odds      

    
    sTicker= ["1"] * 20
    sForwardPE = ["1"] * 20
    sPE = ["1"] *20
    sPS = ["1"] *20
    sPFCF = ["1"] *20
    sDividendPct = ["1"] *20
    sEpsNextY = ["1"] *20
    sEpsPast5Y = ["1"] *20
    sEpsQQ = ["1"] *20
    sSalesQQ = ["1"] *20
    sSharesOutstanding = ["1"] *20
    sProfitMargin = ["1"] *20
    sBeta = ["1"] *20
    sSma20 = ["1"] *20 
    sSma50 = ["1"] *20
    sSma200 = ["1"] *20
    sHigh52W = ["1"] *20
    sLow52W = ["1"] *20
    sRsi = ["1"] *20
    sAvgVol = ["1"] *20
    sPrice = ["1"] *20
    sPriceChangePct = ["1"] *20
    sVolume = ["1"] *20
    sTargetPrice = ["1"] *20
    
    #do it twice for evens and odds
    for i in range (0,len(tickers_Even)):
        sTicker[i] = tickers_Even[i]
        sForwardPE[i] = forwardPE_Even[i]
        sPE[i] = pe_Even[i]
        sPS[i] = ps_Even[i]
        sPFCF[i] = pFCF_Even[i]
        sDividendPct[i] = dividendPct_Even[i]
        sEpsNextY[i] = epsNextY_Even[i]
        sEpsPast5Y[i] = epsPast5Y_Even[i]
        sEpsQQ[i] = epsQQ_Even[i]
        sSalesQQ[i] = salesQQ_Even[i]
        sSharesOutstanding[i] =sharesOutstanding_Even[i]
        sProfitMargin[i] = profitMargin_Even[i]
        sBeta[i] = beta_Even[i]
        sSma20[i] =sma20_Even[i]
        sSma50[i] =sma50_Even[i] 
        sSma200[i] = sma200_Even[i]
        sHigh52W[i] = high52W_Even[i]
        sLow52W[i] = low52W_Even[i]
        sRsi[i] = rsi_Even[i]
        sAvgVol[i] = avgVol_Even[i]
        sPrice[i] =price_Even[i]
        
        sPriceChangePct[i] =priceChangePct_Even[i]
        sPriceChangePct[i]=sPriceChangePct[i].replace("%","")
        
        sVolume[i] = volume_Even[i]
        sTargetPrice[i] = targetPrice_Even[i]
    for i in range (len(tickers_Even),len(tickers_Even)+len(tickers_Odd)):
        sTicker[i]=tickers_Odd[i-len(tickers_Even)]
        sForwardPE[i]=forwardPE_Odd[i-len(tickers_Even)]
        sPE[i] = pe_Odd[i-len(tickers_Even)]
        sPS[i] = ps_Odd[i-len(tickers_Even)]
        sPFCF[i] = pFCF_Odd[i-len(tickers_Even)]
        sDividendPct[i] = dividendPct_Odd[i-len(tickers_Even)]
        sEpsNextY[i] = epsNextY_Odd[i-len(tickers_Even)]
        sEpsPast5Y[i] = epsPast5Y_Odd[i-len(tickers_Even)]
        sEpsQQ[i] = epsQQ_Odd[i-len(tickers_Even)]
        sSalesQQ[i] = salesQQ_Odd[i-len(tickers_Even)]
        sSharesOutstanding[i] = sharesOutstanding_Odd[i-len(tickers_Even)]
        sProfitMargin[i] = profitMargin_Odd[i-len(tickers_Even)]
        sBeta[i] = beta_Odd[i-len(tickers_Even)]
        sSma20[i] =sma20_Odd[i-len(tickers_Even)]
        sSma50[i] =sma50_Odd[i-len(tickers_Even)]
        sSma200[i] =sma200_Odd[i-len(tickers_Even)]         
        sHigh52W[i] = high52W_Odd[i-len(tickers_Even)]   
        sLow52W[i] = low52W_Odd[i-len(tickers_Even)]   
        sRsi[i] = rsi_Odd[i-len(tickers_Even)]   
        sAvgVol[i] =avgVol_Odd[i-len(tickers_Even)]   
        sPrice[i] =price_Odd[i-len(tickers_Even)]   
        
        sPriceChangePct[i] =priceChangePct_Odd[i-len(tickers_Even)]  
        sPriceChangePct[i]=sPriceChangePct[i].replace("%","")
        
        sVolume[i] = volume_Odd[i-len(tickers_Even)]  
        sTargetPrice[i] = targetPrice_Odd[i-len(tickers_Even)]  
    #listOfStockDicts=[]
        
    for i in range (0, len(tickers_Even)+len(tickers_Odd)):
        listOfStockDicts.append({"ticker":sTicker[i], "pe":sPE[i], "forward_PE":sForwardPE[i], "ps":sPS[i], "p_FCF":sPFCF[i], 'dividendPct':sDividendPct[i], 'eps_NextY':sEpsNextY[i], 'epsPast5Y':sEpsPast5Y[i], 'eps_qq':sEpsQQ[i], 'sales_QQ':sSalesQQ[i], 'shares_Outstanding':sSharesOutstanding[i], 'profitMargin':sProfitMargin[i], 'beta':sBeta[i], 'sma_20':sSma20[i], 'sma_50':sSma50[i], 'sma_200':sSma200[i], 'high_52':sHigh52W[i], 'low_52':sLow52W[i], 'rsi_14':sRsi[i], 'avgVol':sAvgVol[i], 'Price':sPrice[i], 'PriceChangePct':sPriceChangePct[i], 'volume':sVolume[i], 'targetPrice':sTargetPrice[i]})
#    FinvizInfoDict = {'Price': price, 'PriceChangePct': priceChangePct, 'beta': beta,'rsi_14':rsi_14, 'shares_Outstanding':sharesOutstand,  'avgVol':avgVol, 'volume':volume, 'targetPrice':targetPrice, 'high_52':high_52, 'low_52':low_52, 'sma_200':sma_200, 'profitMargin':profitMargin, 'eps_NextY':eps_NextY, 'sales_QQ':sales_QQ, 'eps_qq':eps_QQ, 'epsPast5Y':epsPast5Y, 'sma_50':sma_50, 'pe':pe, 'forward_PE':forward_PE,    'ps':ps, 'p_FCF':p_FCF, 'sma_20':sma_20, 'dividendPct':dividendPct}


    
    for item in listOfStockDicts:
        print 'ticker extracted =', item.get("ticker")
    
    
    #return listOfStockDicts
 #   stockticker[i]A,B,C&c=1,7,8,10,13,14,18,19,22,23,24,41,48,52,53,54,57,58,59,63,65,66,67,69'
      

def getFinvizStockListCUMULATIVE(wb):
    sheets = wb.sheetnames
    
    price_ENUM=2
    
    sheet_Price = wb[sheets[price_ENUM]]
    row_num=2
    ticker_col=1
    stockticker = []
    listOfStockDicts=[]
    while (sheet_Price.cell(row=row_num,column=ticker_col).value != None):
        #GET THE 20 STOCKS!
        stockticker.append(str(sheet_Price.cell(row=row_num,column=ticker_col).value))
        if (len(stockticker) == 15):
            FinvizListSite(sheet_Price,stockticker, listOfStockDicts)
            stockticker = []
            
        row_num=row_num+1
        
    print 'listOfStockDicts=',listOfStockDicts
    
    if (len(stockticker) != 0):
        FinvizListSite(stockticker,stockticker, listOfStockDicts)
        
    return listOfStockDicts        
"""
    #currentDate_Column_price = writeTemplate_Price_MAs(sheet_Price, stockLib, curTime)        
    row_num=2
    ticker_col=1        
    for item in listOfStockDicts:
        row_num=2
        while (sheet_Price.cell(row=row_num,column=ticker_col).value != None):
            if (sheet_Price.cell(row=row_num,column=ticker_col).value == item.get("ticker")):
                print 'found ticker', item.get("ticker"), 'in row', row_num
                
                sheet_Price.cell(row=row_num,column=7).value = item.get("ticker")
                
            row_num=row_num+1 
"""




        
#JUST VARS: https://finviz.com/screener.ashx?v=152&c=1,7,8,10,13,14,18,19,22,23,24,41,48,52,53,54,57,58,59,63,66,67,69
#WITH STOCKS AND VARS: https://finviz.com/screener.ashx?v=152&t=A,B,C&c=1,7,8,10,13,14,18,19,22,23,24,41,48,52,53,54,57,58,59,63,65,66,67,69
        
    #print 'stickticker array:', stockticker


def WriteTabs_Price_MAs(wb, stockLib, curTime):
#These three tabs have the same template.  Should still compare the ticker list to make sure no errors occured
    #go by tab num
    green = Font(color=GREEN)
    red = Font(color=RED)
    
    price_ENUM=2
    Beta_ENUM=3
    Rsi14_ENUM=4
    sharesOutstanding_ENUM=5
    avgVol_ENUM=6
    volume_ENUM=7
    targetPrice_ENUM=8
    high52W_ENUM=9
    low52W_ENUM=10
    sma20_ENUM=11
    sma50_ENUM=12
    sma200_ENUM=13
    profitMargin_ENUM=14
    eps_NextY_ENUM=15
    sales_QQ_ENUM=16
    eps_qq_ENUM=17
    epsPast5Y_ENUM=18
    pe_ENUM=19
    forwardPE_ENUM=20
    ps_ENUM=21
    p_FCF_ENUM=22
    dividendPct_ENUM=23
    
    sheets = wb.sheetnames
    print 'sheets2: ' ,sheets
    sheet_Price = wb[sheets[price_ENUM]]
    sheet_Beta = wb[sheets[Beta_ENUM]]
    sheet_RSI14 = wb[sheets[Rsi14_ENUM]]
    sheet_SharesOutstanding=wb[sheets[sharesOutstanding_ENUM]]
    sheet_AvgVolume = wb[sheets[avgVol_ENUM]]
    sheet_Volume = wb[sheets[volume_ENUM]]
    sheet_TargetPrice = wb[sheets[targetPrice_ENUM]]
    sheet_High52W = wb[sheets[high52W_ENUM]]
    sheet_Low52W = wb[sheets[low52W_ENUM]]
    sheet_Sma20 = wb[sheets[sma20_ENUM]]
    sheet_Sma50 = wb[sheets[sma50_ENUM]]
    sheet_Sma200 = wb[sheets[sma200_ENUM]] 
    sheet_profitMargin = wb[sheets[profitMargin_ENUM]]
    sheet_eps_NextY = wb[sheets[eps_NextY_ENUM]]
    sheet_sales_QQ = wb[sheets[sales_QQ_ENUM]] 
    sheet_eps_qq=wb[sheets[eps_qq_ENUM]]
    sheet_epsPast5Y=wb[sheets[epsPast5Y_ENUM]]
    sheet_pe=wb[sheets[pe_ENUM]]
    sheet_forwardPE=wb[sheets[forwardPE_ENUM]]
    sheet_ps=wb[sheets[ps_ENUM]]
    sheet_p_FCF=wb[sheets[p_FCF_ENUM]]
    sheet_dividendPct=wb[sheets[dividendPct_ENUM]]
    
    currentDate_Column_price = writeTemplate_Price_MAs(sheet_Price, stockLib, curTime)
    currentDate_Column_Beta = writeTemplate_Price_MAs(sheet_Beta, stockLib, curTime)
    currentDate_Column_RSI14 = writeTemplate_Price_MAs(sheet_RSI14, stockLib, curTime)
    writeTemplate_Price_MAs(sheet_SharesOutstanding, stockLib, curTime)
    writeTemplate_Price_MAs(sheet_AvgVolume, stockLib, curTime)
    writeTemplate_Price_MAs(sheet_Volume, stockLib, curTime)
    writeTemplate_Price_MAs(sheet_TargetPrice, stockLib, curTime)
    writeTemplate_Price_MAs(sheet_High52W, stockLib, curTime)
    writeTemplate_Price_MAs(sheet_Low52W, stockLib, curTime)
    writeTemplate_Price_MAs(sheet_Sma20, stockLib, curTime)
    writeTemplate_Price_MAs(sheet_Sma50, stockLib, curTime)
    writeTemplate_Price_MAs(sheet_Sma200, stockLib, curTime)
    writeTemplate_Price_MAs(sheet_profitMargin, stockLib, curTime)
    writeTemplate_Price_MAs(sheet_eps_NextY, stockLib, curTime)
    writeTemplate_Price_MAs(sheet_sales_QQ, stockLib, curTime)
    writeTemplate_Price_MAs(sheet_eps_qq, stockLib, curTime)
    writeTemplate_Price_MAs(sheet_epsPast5Y, stockLib, curTime)
    writeTemplate_Price_MAs(sheet_pe, stockLib, curTime)    
    writeTemplate_Price_MAs(sheet_forwardPE, stockLib, curTime)
    writeTemplate_Price_MAs(sheet_ps, stockLib, curTime)
    writeTemplate_Price_MAs(sheet_p_FCF, stockLib, curTime)
    writeTemplate_Price_MAs(sheet_dividendPct, stockLib, curTime)
    
    #make sure dates are the same!
    if((currentDate_Column_price == currentDate_Column_Beta) and (currentDate_Column_price == currentDate_Column_RSI14)):
        currentDate_column = currentDate_Column_price
        print 'All column dates are the same.  Good! currentDate_column = ', currentDate_column
    else:
        print 'Column Dates are not the same, error!'
        print 'currentDate_Column_price = ', currentDate_Column_price
        print 'currentDate_Column_Beta = ', currentDate_Column_Beta
        print 'currentDate_Column_RSI14 = ', currentDate_Column_RSI14
        
    cumulative_StockLib = {}
    listOfStockDicts = getFinvizStockListCUMULATIVE(wb)
    stock_col=1
    stockMatch=True
    #make sure stock lists are the same!
    row_num=2

    row_num=2
    ticker_col=1            
    for item in listOfStockDicts:
        row_num=2
        while (sheet_Price.cell(row=row_num,column=ticker_col).value != None):
            #match up the ticker from the dict with the excel
            if (sheet_Price.cell(row=row_num,column=ticker_col).value == item.get("ticker")):
                #make sure the tickers are the same
                if ((sheet_Price.cell(row=row_num,column=stock_col).value == sheet_Beta.cell(row=row_num,column=stock_col).value) and  (sheet_Price.cell(row=row_num,column=stock_col).value == sheet_RSI14.cell(row=row_num,column=stock_col).value)):
                    #at this point, row_num is the exact row number that the stock ticker (matched) is on
                    print 'found ticker', item.get("ticker"), 'in row', row_num
        #listOfStockDicts.append({"ticker":sTicker[i], "pe":sPE[i], "forward_PE":sForwardPE[i], "ps":sPS[i], "p_FCF":sPFCF[i], 'dividendPct':sDividendPct[i], 'eps_NextY':sEpsNextY[i], 'epsPast5Y':sEpsPast5Y[i], 'eps_qq':sEpsQQ[i], 'sales_QQ':sSalesQQ[i], 'shares_Outstanding':sSharesOutstanding[i], 'profitMargin':sProfitMargin[i], 'beta':sBeta[i], 'sma_20':sSma20[i], 'sma_50':sSma50[i], 'sma_200':sSma200[i], 'high_52':sHigh52W[i], 'low_52':sLow52W[i], 'rsi_14':sRsi[i], 'avgVol':sAvgVol[i], 'Price':sPrice[i], 'PriceChangePct':sPriceChangePct[i], 'volume':sVolume, 'targetPrice':sTargetPrice})
           
                    #add in Price
                    stockPriceTemplate = str(item.get("Price")) + " (" + str(item.get("PriceChangePct")) + "%)"
                    sheet_Price.cell(row=row_num,column=currentDate_column).value = stockPriceTemplate
                    if (float(item.get("PriceChangePct")) >= 0):
                        sheet_Price.cell(row=row_num,column=currentDate_column).font = green
                    else:
                        sheet_Price.cell(row=row_num,column=currentDate_column).font = red

                    #Second Add in BETA
                    sheet_Beta.cell(row=row_num,column=currentDate_column).value = item.get("beta")
                    #third Add in RSI 14
                    sheet_RSI14.cell(row=row_num,column=currentDate_column).value = float(item.get("rsi_14"))

                    #Add color if over /under moving averages
                    if(float(item.get("rsi_14")) < 30 ):
                        sheet_RSI14.cell(row=row_num,column=currentDate_column).font = green
                    elif(float(item.get("rsi_14")) > 70):
                        sheet_RSI14.cell(row=row_num,column=currentDate_column).font = red
                    else:
                        sheet_RSI14.cell(row=row_num,column=currentDate_column).font = None
                        
                                   
                    #add in SharesOutstanding
                    sheet_SharesOutstanding.cell(row=row_num,column=currentDate_column).value = item.get("shares_Outstanding")

                    #add in AvgVolume
                    sheet_AvgVolume.cell(row=row_num,column=currentDate_column).value = item.get("avgVol")
                    
                    #add in Volume
                    sheet_Volume.cell(row=row_num,column=currentDate_column).value = item.get("volume")            
                    
                    #add in target price
                    sheet_TargetPrice.cell(row=row_num,column=currentDate_column).value = item.get("targetPrice")    
                    
                    #add in Dist from 52 week high
                    sheet_High52W.cell(row=row_num,column=currentDate_column).value = item.get("high_52")   
                    
                    #add in dist from 52W Low
                    sheet_Low52W.cell(row=row_num,column=currentDate_column).value = item.get("low_52")   
            
                    #add in dist from sma20
                    sheet_Sma20.cell(row=row_num,column=currentDate_column).value = item.get("sma_20")       
         
                    #add in dist from sma50
                    sheet_Sma50.cell(row=row_num,column=currentDate_column).value = item.get("sma_50")     
                    
                    #add in dist from sma200
                    sheet_Sma200.cell(row=row_num,column=currentDate_column).value = item.get("sma_200")     

                    #add in profit margin
                    sheet_profitMargin.cell(row=row_num,column=currentDate_column).value = item.get("profitMargin")   
                    
                    #eps next Y
                    sheet_eps_NextY.cell(row=row_num,column=currentDate_column).value = item.get("eps_NextY")   
                    
                    #Sales QQ
                    sheet_sales_QQ.cell(row=row_num,column=currentDate_column).value = item.get("sales_QQ")   

                    #eps_qq
                    sheet_eps_qq.cell(row=row_num,column=currentDate_column).value = item.get("eps_qq")   

                    #eps Past 5 Years
                    sheet_epsPast5Y.cell(row=row_num,column=currentDate_column).value = item.get("epsPast5Y")   

                    #PE
                    sheet_pe.cell(row=row_num,column=currentDate_column).value = item.get("pe")   

                    #Forware PE
                    sheet_forwardPE.cell(row=row_num,column=currentDate_column).value = item.get("forward_PE")   
                    
                    #PS
                    sheet_ps.cell(row=row_num,column=currentDate_column).value = item.get("ps")   
                    
                    #p_FCF
                    sheet_p_FCF.cell(row=row_num,column=currentDate_column).value = item.get("p_FCF")   
                    
                    #dividend pct
                    sheet_dividendPct.cell(row=row_num,column=currentDate_column).value = item.get("dividendPct")   
                    
                    #add to a dictionary.  Maybe will be useful? 
                    cumulative_StockLib[item.get("ticker")] = [float(item.get("Price"))]                    
                
            row_num=row_num+1
                
    return cumulative_StockLib
            
            #add to a dictionary.  Maybe will be useful? 
            #cumulative_StockLib[stockTicker] = [float(FinvizDict.get("Price"))]    
        
    
"""    
    while (sheet_Price.cell(row=row_num,column=stock_col).value != None):
        #print 'row_num= ', row_num
        if ((sheet_Price.cell(row=row_num,column=stock_col).value == sheet_Beta.cell(row=row_num,column=stock_col).value) and  (sheet_Price.cell(row=row_num,column=stock_col).value == sheet_RSI14.cell(row=row_num,column=stock_col).value)):
            #cumulativeStockLib{stockticker:price, 50dayma,200dayma}          
            
""
            stockTicker=str(sheet_Price.cell(row=row_num,column=stock_col).value)
            #Data from yahoo SATISTICS PAGE
            #treeYahoo = GetYahoo_StatisticsPage(stockTicker)     
            print 'on stock : ', stockTicker

            FinvizDict={}
            finvizTree = GetFinviz_StockPage(stockTicker)
            FinvizDict = GetFinvizStockINFO(finvizTree)
            print 'FinvizDict=', FinvizDict
#    FinvizInfoDict = {'Price': price, 'PriceChangePct': priceChangePct, 'beta': beta,'rsi_14':rsi_14, 'shares_Outstanding':sharesOutstand,  'avgVol':avgVol, 'volume':volume, 'targetPrice':targetPrice, 'high_52':high_52, 'low_52':low_52, 'sma_200':sma_200, 'profitMargin':profitMargin, 'eps_NextY':eps_NextY, 'sales_QQ':sales_QQ, 'eps_qq':eps_QQ, 'epsPast5Y':epsPast5Y, 'sma_50':sma_50, 'pe':pe, 'forward_PE':forward_PE,    'ps':ps, 'p_FCF':p_FCF, 'sma_20':sma_20, 'dividendPct':dividendPct}
            
            #FIRST add in PRICE in form "190 (-.5%)"   
            print 'get!', FinvizDict.get('rsi_14')
            stockPriceTemplate = str(FinvizDict.get("Price")) + " (" + str(FinvizDict.get("PriceChangePct")) + "%)"
            
            sheet_Price.cell(row=row_num,column=currentDate_column).value = stockPriceTemplate
            print 'float(FinvizDict.get("PriceChangePct"))=', FinvizDict.get("PriceChangePct")
            if (float(FinvizDict.get("PriceChangePct")) >= 0):
                sheet_Price.cell(row=row_num,column=currentDate_column).font = green
            else:
                sheet_Price.cell(row=row_num,column=currentDate_column).font = red
                
            #Second Add in BETA
            sheet_Beta.cell(row=row_num,column=currentDate_column).value = FinvizDict.get("beta")
                            
            #third Add in RSI 14
            sheet_RSI14.cell(row=row_num,column=currentDate_column).value = float(FinvizDict.get("rsi_14"))

            #Add color if over /under moving averages
            if(float(FinvizDict.get("rsi_14")) < 30 ):
                sheet_RSI14.cell(row=row_num,column=currentDate_column).font = green
            elif(float(FinvizDict.get("rsi_14")) > 70):
                sheet_RSI14.cell(row=row_num,column=currentDate_column).font = red
            else:
                sheet_RSI14.cell(row=row_num,column=currentDate_column).font = None
                
                           
            #add in SharesOutstanding
            sheet_SharesOutstanding.cell(row=row_num,column=currentDate_column).value = FinvizDict.get("shares_Outstanding")

            #add in AvgVolume
            sheet_AvgVolume.cell(row=row_num,column=currentDate_column).value = FinvizDict.get("avgVol")
            
            #add in Volume
            sheet_Volume.cell(row=row_num,column=currentDate_column).value = FinvizDict.get("volume")            
            
            #add in target price
            sheet_TargetPrice.cell(row=row_num,column=currentDate_column).value = FinvizDict.get("targetPrice")    
            
            #add in Dist from 52 week high
            sheet_High52W.cell(row=row_num,column=currentDate_column).value = FinvizDict.get("high_52")   
            
            #add in dist from 52W Low
            sheet_Low52W.cell(row=row_num,column=currentDate_column).value = FinvizDict.get("low_52")   
    
            #add in dist from sma20
            sheet_Sma20.cell(row=row_num,column=currentDate_column).value = FinvizDict.get("sma_20")       
 
            #add in dist from sma50
            sheet_Sma50.cell(row=row_num,column=currentDate_column).value = FinvizDict.get("sma_50")     
            
            #add in dist from sma200
            sheet_Sma200.cell(row=row_num,column=currentDate_column).value = FinvizDict.get("sma_200")     

            #add in profit margin
            sheet_profitMargin.cell(row=row_num,column=currentDate_column).value = FinvizDict.get("profitMargin")   
            
            #eps next Y
            sheet_eps_NextY.cell(row=row_num,column=currentDate_column).value = FinvizDict.get("eps_NextY")   
            
            #Sales QQ
            sheet_sales_QQ.cell(row=row_num,column=currentDate_column).value = FinvizDict.get("sales_QQ")   

            #eps_qq
            sheet_eps_qq.cell(row=row_num,column=currentDate_column).value = FinvizDict.get("eps_qq")   

            #eps Past 5 Years
            sheet_epsPast5Y.cell(row=row_num,column=currentDate_column).value = FinvizDict.get("epsPast5Y")   

            #PE
            sheet_pe.cell(row=row_num,column=currentDate_column).value = FinvizDict.get("pe")   

            #Forware PE
            sheet_forwardPE.cell(row=row_num,column=currentDate_column).value = FinvizDict.get("forward_PE")   
            
            #PS
            sheet_ps.cell(row=row_num,column=currentDate_column).value = FinvizDict.get("ps")   
            
            #p_FCF
            sheet_p_FCF.cell(row=row_num,column=currentDate_column).value = FinvizDict.get("p_FCF")   
            
            #dividend pct
            sheet_dividendPct.cell(row=row_num,column=currentDate_column).value = FinvizDict.get("dividendPct")   
 
            #add to a dictionary.  Maybe will be useful? 
            cumulative_StockLib[stockTicker] = [float(FinvizDict.get("Price"))]
        else:
            stockMatch=False
            break
        row_num=row_num+1
    if(stockMatch == True):
        print 'stocks match, continue'
    else:
        print 'ERROR: stocks Dont Match!'
    
    print 'cumulative_StockLibrary: ',cumulative_StockLib
"""    

    #check if stock exists
#    for key in stockLib.keys():

   
def WriteTab_Options(wb, stockLib, curTime):
    sheets = wb.sheetnames
    print 'sheets1: ' ,sheets
    sheet_options = wb[sheets[5]]
    
    #See if we already have options info on that stock, if not add it in
    stockCol_num=1
    for key in stockLib.keys():
        row_num=1
        found = 0
        while ((sheet_options.cell(row=row_num, column=stockCol_num).value != None) and (found  == 0)):
            if (key == sheet_options.cell(row=row_num, column=stockCol_num).value):
                found = 1
            else:
                found = 0
            row_num=row_num+16
        if (found == 1):
            print 'found key: ', key
        else:
            print 'could not find key: ', key
            print 'row_num = ', row_num
            newStockRow=row_num
            #insert stock info!
            sheet_options.cell(row=newStockRow, column=stockCol_num).value = key
            tree = GetBarChart_EarningsPage(key)
            earningsDates=[]
            earningsDates = GetBarChart_Dates(tree)
            if earningsDates: 
                earningsData=[]
                earningsData = GetBarChart_Table(tree)
            
                #First Table
                sheet_options.cell(row=newStockRow, column=stockCol_num + 1).value = 'Earnings History - Surprises'            
                sheet_options.cell(row=newStockRow + 2, column=stockCol_num + 1).value = 'Reported'
                sheet_options.cell(row=newStockRow + 3, column=stockCol_num + 1).value = 'Estimate'
                sheet_options.cell(row=newStockRow + 4, column=stockCol_num + 1).value = 'Difference'
                sheet_options.cell(row=newStockRow + 5, column=stockCol_num + 1).value = 'Surprise'
            
                #Earnings Dates Tbl 1
                print 'EARNINGDATES= ',earningsDates
                sheet_options.cell(row=newStockRow + 1, column=stockCol_num + 2).value = earningsDates[0] + " " + earningsDates[1]
                sheet_options.cell(row=newStockRow + 1, column=stockCol_num + 3).value = earningsDates[2] + " " + earningsDates[3]
                sheet_options.cell(row=newStockRow + 1, column=stockCol_num + 4).value = earningsDates[4] + " " + earningsDates[5]
                sheet_options.cell(row=newStockRow + 1, column=stockCol_num + 5).value = earningsDates[6] + " " + earningsDates[7]
                
                #Data Tbl 1
                sheet_options.cell(row=newStockRow + 2, column=stockCol_num + 2).value = earningsData[2]
                sheet_options.cell(row=newStockRow + 2, column=stockCol_num + 3).value = earningsData[3]
                sheet_options.cell(row=newStockRow + 2, column=stockCol_num + 4).value = earningsData[4]
                sheet_options.cell(row=newStockRow + 2, column=stockCol_num + 5).value = earningsData[5]
                
                sheet_options.cell(row=newStockRow + 3, column=stockCol_num + 2).value = earningsData[8]
                sheet_options.cell(row=newStockRow + 3, column=stockCol_num + 3).value = earningsData[9]
                sheet_options.cell(row=newStockRow + 3, column=stockCol_num + 4).value = earningsData[10]
                sheet_options.cell(row=newStockRow + 3, column=stockCol_num + 5).value = earningsData[11]
                
                sheet_options.cell(row=newStockRow + 4, column=stockCol_num + 2).value = earningsData[14]
                sheet_options.cell(row=newStockRow + 4, column=stockCol_num + 3).value = earningsData[15]
                sheet_options.cell(row=newStockRow + 4, column=stockCol_num + 4).value = earningsData[16]
                sheet_options.cell(row=newStockRow + 4, column=stockCol_num + 5).value = earningsData[17]  

                sheet_options.cell(row=newStockRow + 5, column=stockCol_num + 2).value = earningsData[20]
                sheet_options.cell(row=newStockRow + 5, column=stockCol_num + 3).value = earningsData[21]
                sheet_options.cell(row=newStockRow + 5, column=stockCol_num + 4).value = earningsData[22]
                sheet_options.cell(row=newStockRow + 5, column=stockCol_num + 5).value = earningsData[23]                
                
                #Second Table
                sheet_options.cell(row=newStockRow + 7, column=stockCol_num + 1).value = 'Earnings Estimates'            
                sheet_options.cell(row=newStockRow + 9, column=stockCol_num + 1).value = 'Average Estimate'
                sheet_options.cell(row=newStockRow + 10, column=stockCol_num + 1).value = 'Number of Estimates'
                sheet_options.cell(row=newStockRow + 11, column=stockCol_num + 1).value = 'High Estimate'
                sheet_options.cell(row=newStockRow + 12, column=stockCol_num + 1).value = 'Low Estimate'
                sheet_options.cell(row=newStockRow + 13, column=stockCol_num + 1).value = 'Prior Year'        
                sheet_options.cell(row=newStockRow + 14, column=stockCol_num + 1).value = 'Growth Rate Est. YoY'        

                #earnings dates tbl 2
                sheet_options.cell(row=newStockRow + 8, column=stockCol_num + 2).value = earningsDates[8] + " " + earningsDates[9]
                sheet_options.cell(row=newStockRow + 8, column=stockCol_num + 3).value = earningsDates[10] + " " + earningsDates[11]
                sheet_options.cell(row=newStockRow + 8, column=stockCol_num + 4).value = earningsDates[12] + " " + earningsDates[13]
                sheet_options.cell(row=newStockRow + 8, column=stockCol_num + 5).value = earningsDates[14] + " " + earningsDates[15]
                
                #data Tbl 2
                sheet_options.cell(row=newStockRow + 9, column=stockCol_num + 2).value = earningsData[26]
                sheet_options.cell(row=newStockRow + 9, column=stockCol_num + 3).value = earningsData[27]
                sheet_options.cell(row=newStockRow + 9, column=stockCol_num + 4).value = earningsData[28]
                sheet_options.cell(row=newStockRow + 9, column=stockCol_num + 5).value = earningsData[29]
                
                sheet_options.cell(row=newStockRow + 10, column=stockCol_num + 2).value = earningsData[32]
                sheet_options.cell(row=newStockRow + 10, column=stockCol_num + 3).value = earningsData[33]
                sheet_options.cell(row=newStockRow + 10, column=stockCol_num + 4).value = earningsData[34]
                sheet_options.cell(row=newStockRow + 10, column=stockCol_num + 5).value = earningsData[35]                
                
                sheet_options.cell(row=newStockRow + 11, column=stockCol_num + 2).value = earningsData[38]
                sheet_options.cell(row=newStockRow + 11, column=stockCol_num + 3).value = earningsData[39]
                sheet_options.cell(row=newStockRow + 11, column=stockCol_num + 4).value = earningsData[40]
                sheet_options.cell(row=newStockRow + 11, column=stockCol_num + 5).value = earningsData[41]               

                sheet_options.cell(row=newStockRow + 12, column=stockCol_num + 2).value = earningsData[44]
                sheet_options.cell(row=newStockRow + 12, column=stockCol_num + 3).value = earningsData[45]
                sheet_options.cell(row=newStockRow + 12, column=stockCol_num + 4).value = earningsData[46]
                sheet_options.cell(row=newStockRow + 12, column=stockCol_num + 5).value = earningsData[47]     

                sheet_options.cell(row=newStockRow + 13, column=stockCol_num + 2).value = earningsData[50]
                sheet_options.cell(row=newStockRow + 13, column=stockCol_num + 3).value = earningsData[51]
                sheet_options.cell(row=newStockRow + 13, column=stockCol_num + 4).value = earningsData[52]
                sheet_options.cell(row=newStockRow + 13, column=stockCol_num + 5).value = earningsData[53]        

                sheet_options.cell(row=newStockRow + 14, column=stockCol_num + 2).value = earningsData[57]
                sheet_options.cell(row=newStockRow + 14, column=stockCol_num + 3).value = earningsData[58]
                sheet_options.cell(row=newStockRow + 14, column=stockCol_num + 4).value = earningsData[59]
                sheet_options.cell(row=newStockRow + 14, column=stockCol_num + 5).value = earningsData[60]                     
            else:
                print 'stock not on list: ', key
                sheet_options.cell(row=newStockRow, column=stockCol_num + 1).value = 'STOCK NOT ON LIST!'            

            
            row_num=row_num+16

def WriteTotalChangeSinceInception(wb, cumulative_StockLib, curTime):
    sheets = wb.sheetnames
    #print 'sheets1: ' ,sheets
    sheet1 = wb[sheets[1]]
    dateStr=str(curTime.month)+"/"+str(curTime.day)+"/"+str(curTime.year)     
    
    green = Font(color=GREEN)
    red = Font(color=RED)
    
    col_num=1
    row_num = 1
    
    #get the last column to put the totals
    while (sheet1.cell(row=row_num,column=col_num).value != None and sheet1.cell(row=row_num,column=col_num).value != '% Change Since Inception'):
        col_num = col_num +1
    totals_col = col_num
    
    sheet1.cell(row=1,column=col_num).value = "% Change Since Inception"
    
    price_ENUM=0
    row_num = 1
    #column=1 is the tickers
    while (sheet1.cell(row=row_num,column=1).value != None):
        for key in cumulative_StockLib.keys():
            if (sheet1.cell(row=row_num,column=1).value == key):
                i=0
                for item in cumulative_StockLib[key]: 
                    if (i==price_ENUM):
                        #print 'current price=', item
                        price_Change = item - sheet1.cell(row=row_num,column=2).value
                        pctChange = round(float(100*(price_Change/sheet1.cell(row=row_num,column=2).value)),2)
                        #print 'pctChange =', pctChange
                        sheet1.cell(row=row_num,column=totals_col).value = pctChange
                        if (pctChange > 0):
                            sheet1.cell(row=row_num,column=totals_col).font = green
                        else:
                            sheet1.cell(row=row_num,column=totals_col).font = red
                    i=i+1
        row_num=row_num+1
        
def WriteOverallStats(wb, cumulative_StockLib, curTime):
    sheets = wb.sheetnames
    #print 'sheets1: ' ,sheets
    sheet1 = wb[sheets[1]]
    dateStr=str(curTime.month)+"/"+str(curTime.day)+"/"+str(curTime.year)     
    
    green = Font(color=GREEN)
    red = Font(color=RED)
    
    col_num=1    
    while (sheet1.cell(row=1,column=col_num).value != '% Change Since Inception'):
        col_num = col_num +1
    totals_col = col_num
    
    #start at 2 because 1 is the '%change since inception'
    row_num = 2
    total_pctChange=0
    while (sheet1.cell(row=row_num,column=totals_col).value != None):
        total_pctChange=total_pctChange+sheet1.cell(row=row_num,column=totals_col).value
        row_num=row_num+1
        
    sheet1['P7'] = "Total Sum:"
    sheet1['Q7'] = str(total_pctChange) + "%"

def WriteFidelityStats(wb, cumulative_StockLib, curTime):
    sheets = wb.sheetnames
    #print 'sheets1: ' ,sheets
    sheet1 = wb[sheets[1]]
    dateStr=str(curTime.month)+"/"+str(curTime.day)+"/"+str(curTime.year)     
    
    green = Font(color=GREEN)
    red = Font(color=RED)

    col_num=1    
    while (sheet1.cell(row=1,column=col_num).value != 'Buy Sell Ratio'):
        col_num = col_num +1
    buysell_Ratio = col_num    
    
    col_num=1    
    while (sheet1.cell(row=1,column=col_num).value != '% Change Since Inception'):
        col_num = col_num +1
    totals_col = col_num
    
    row_num = 2
    greater_75_pctChange=0
    lower_30_pctChange=0
    btwn_35_65_pctChange=0
    while (sheet1.cell(row=row_num,column=totals_col).value != None):
        if(sheet1.cell(row=row_num,column=buysell_Ratio).value >= .75):
            print 'over .75 stock : ', sheet1.cell(row=row_num,column=1).value
            print 'totals col = ', totals_col
            print 'over .75 stock totals col = ', sheet1.cell(row=row_num,column=totals_col).value
            greater_75_pctChange=greater_75_pctChange+sheet1.cell(row=row_num,column=totals_col).value
        if (sheet1.cell(row=row_num,column=buysell_Ratio).value >= .35 and sheet1.cell(row=row_num,column=buysell_Ratio).value <= .65):
            btwn_35_65_pctChange=btwn_35_65_pctChange+sheet1.cell(row=row_num,column=totals_col).value
        if (sheet1.cell(row=row_num,column=buysell_Ratio).value <= .30):
            lower_30_pctChange=lower_30_pctChange+sheet1.cell(row=row_num,column=totals_col).value
        row_num=row_num+1
    
    sheet1['P8'] = ".75+ Ratio"
    sheet1['Q8'] = str(greater_75_pctChange) + "%"
    
    sheet1['P9'] = ".35-.65 Ratio"
    sheet1['Q9'] = str(btwn_35_65_pctChange) + "%"    
        
    sheet1['P10'] = ".30- Ratio"
    sheet1['Q10'] = str(lower_30_pctChange) + "%"
    
def csvWriter(stockLib, listName):
    curTime = datetime.datetime.now()
    
    
    backslash='\abc'
    print 'backslash=', backslash
    #workbook = xlsxwriter.Workbook('demo.xlsx')
    filepath = os.path.join(listName, listName +'.xlsx')
    #os.path.join(r"C:\mypath", "subfolder")
    #filepath=listName + '/' + listName + '.xlsx'
    filepath=listName+'/'+listName+'.xlsx'
    print 'filepath =',filepath
    #filepath='demo_2.xlsx'
    wb=load_workbook(filepath)
    
    getFinvizStockListCUMULATIVE(wb)
    
    
#    WriteTab_DailyStockList(workbook, stockLib, curTime)
    WriteTab_DailyStockList(wb, stockLib, curTime)
    
    WriteTab_CumulativeStockList(wb, stockLib, curTime)
    
    cumulative_StockLib = {}
    cumulative_StockLib = WriteTabs_Price_MAs(wb, stockLib, curTime)
        
    #NOTE: THIS SHOULD BE THE SAME AS stockLib, not cumulative_StockLib (Only for testing!)
    #lets not do the options tab for finviz
    #WriteTab_Options(wb, stockLib, curTime)

    WriteTotalChangeSinceInception(wb, cumulative_StockLib, curTime)
    
    #WriteOverallStats(wb, cumulative_StockLib, curTime)

    #WriteFidelityStats(wb, cumulative_StockLib, curTime)

    
    wb.save(filepath)
    #workbook.close()

    
   
    
def main():

    #Stock Lib is a dictionary.  It is how we return all of the values from this function.  Keep in mind it does not have labels.
    #Returns as #{{ticker1: price1, buy#1,sell#1,ratio#1},{ticker2: price2, buy#2,sell#2,ratio#2}, {3....}... }
    #Will need to figure out how to add yahoo functions to it as well.  Also need to figure out how to make it look nice in excel
    #earningDates=[]
    #tree = GetBarChart_EarningsPage('EA')
    #EaEarnings = []
    #EaEarnings = GetBarChart_Table(tree)
    #for i in range (0,len(EaEarnings)):
    #    print 'EaEarnings[i]',i, '=', EaEarnings[i]
    #print 'earning dates= ', earningDates

    #print 'earning dates[0] = ', earningDates[0]
    #print 'earning dates[1] = ', earningDates[1]
    #abc = earningDates[0] + " " + earningDates[1]
    #print 'abc = ', abc
    listName=sys.argv[1]
    print 'listName=',listName 

    #FinvizDict={}
    #finvizTree = GetFinviz_StockPage("ACCO")
    #FinvizDict = GetFinvizStockINFO_test(finvizTree)
    #beta,price = GetFinvizStockINFO_Price_Beta(finvizTree)
    #print 'beta= ', beta
    #print 'price= ', price
    
    #priceChangePct=GetFinvizStockINFO_PriceChangePct(finvizTree)
    #print 'priceChangePct=', priceChangePct
    
    if listName == 'Finviz_TLSupport_Oversold40_InvHammer':
        web = 'https://finviz.com/screener.ashx?v=152&f=geo_usa,ind_stocksonly,ta_candlestick_ih,ta_pattern_tlsupport,ta_rsi_os40&ft=3&c=1,2,3,4,6,48,59,65,66,67'
    elif listName == 'Finviz_WedgeStrng_Oversold40_AvgtrueRngUndr25':
        web = 'https://finviz.com/screener.ashx?v=152&f=geo_usa,ind_stocksonly,ta_averagetruerange_u2.5,ta_pattern_wedge2,ta_rsi_os40&ft=3&c=1,2,3,4,6,48,59,65,66,67'
    elif listName == 'Finviz_MajorNews':
        web = 'https://finviz.com/screener.ashx?v=152&s=n_majornews&f=geo_usa,ind_stocksonly&ft=3&c=1,2,3,4,6,48,59,65,66,67'
    elif listName == 'Finviz_Upgrades':
        web = 'https://finviz.com/screener.ashx?v=152&s=n_upgrades&f=geo_usa,ind_stocksonly&ft=3&c=1,2,3,4,6,48,59,65,66,67'
    elif listName == 'Finviz_TopGainers_PEless15':
        web = 'https://finviz.com/screener.ashx?v=152&s=ta_topgainers&f=fa_fpe_low,geo_usa,ind_stocksonly&ft=4&c=1,2,3,4,6,48,59,65,66,67'
    elif listName == 'Finviz_Sma20Bounce':
        web = 'https://finviz.com/screener.ashx?v=152&f=sh_avgvol_o2000,sh_curvol_o2000,sh_price_10to50,sh_relvol_o1,ta_sma20_pa,ta_sma50_pb&ft=4&o=industry&c=1,2,3,4,6,48,59,65,66,67'
    elif listName == 'Finviz_OversoldEarningsMo':
        web = 'https://finviz.com/screener.ashx?v=152&f=cap_smallover,earningsdate_thismonth,fa_epsqoq_o15,fa_grossmargin_o20,sh_avgvol_o750,sh_curvol_o1000,ta_perf_52w10o,ta_rsi_nob50&ft=4&c=1,2,3,4,6,48,59,65,66,67'
    elif listName == 'Finviz_Rsi30Reversal':
        web = 'https://finviz.com/screener.ashx?v=152&f=sh_price_o5,sh_relvol_o2,ta_change_u,ta_rsi_os30&ft=4&o=price&c=1,2,3,4,6,48,59,65,66,67'
    elif listName == 'Finviz_SmallCap_InstutionalTrans50Pct_30-45RSI_30SalesGrwth':
        web = 'https://finviz.com/screener.ashx?v=152&f=cap_small,fa_salesqoq_o30,sh_insttrans_o50,ta_rsi_30to45&ft=4&c=1,2,3,4,6,48,59,65,66,67' #https://www.guerillastocktrading.com/lessons/finviz-stock-screener-killer-setups/
    else:
        print 'could not find listName: ', listName
        #web='https://finviz.com/screener.ashx?v=150&f=geo_usa,ind_stocksonly,ta_candlestick_ih,ta_pattern_tlsupport,ta_rsi_os40&ft=3'
        #web='https://finviz.com/screener.ashx?v=111&f=geo_usa,ind_stocksonly,ta_candlestick_ih,ta_pattern_tlsupport,ta_rsi_os40&ft=3'
    print 'web=',web
        
    stockLib = {}
    stockLib = FinvizStockParse(web)
#    stockLib = StockParse()
    
#    for key in stockLib.keys():
#        i=0
#        for item in stockLib[key]: 
#            if (i==stockLib_sBeta):
#                if (item > 5 and item != 'N/A'):
#                    stockLib.pop(key,None)
#                    print 'removing key: ', key
#            i=i+1
    print 'stockLib in MAIN ', stockLib    
    
    csvWriter(stockLib, listName)
    
    print 'Ticker: Price, PriceChange, price change %, buy amount, sell amount, ratio, beta'




if __name__ == '__main__':
    main()